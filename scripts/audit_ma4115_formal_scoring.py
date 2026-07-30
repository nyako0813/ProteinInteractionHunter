#!/usr/bin/env python3
"""Independent real-data audit for the MA_4115 formal scoring pilot."""

from __future__ import annotations

import argparse
import csv
import json
import math
import statistics
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

PRECISION = 12
TIE_PRECISION = 8
OUTPUT_SCALE = 100.0
MINIMUM_WEIGHT = 1.0
MINIMUM_CATEGORIES = 2
WEIGHTS = {
    "genome_context": 1.0,
    "operon_proxy": 1.0,
    "domain_pair": 1.0,
    "functional_complementarity": 1.0,
    "localization": 0.5,
    "orthology": 0.75,
    "phylogenetic_profile": 1.0,
    "fusion": 1.5,
    "known_interactions": 1.5,
}
CATEGORIES = {
    "genome_context": "genomic_context",
    "operon_proxy": "genomic_context",
    "domain_pair": "functional_annotation",
    "functional_complementarity": "functional_annotation",
    "localization": "cellular_compatibility",
    "orthology": "evolutionary",
    "phylogenetic_profile": "evolutionary",
    "fusion": "direct_interaction",
    "known_interactions": "direct_interaction",
}
CAPS = {
    "genomic_context": 1.5,
    "functional_annotation": 1.5,
    "cellular_compatibility": 0.5,
    "evolutionary": 2.0,
    "direct_interaction": 2.0,
}
COMPONENT_ORDER = tuple(WEIGHTS)
DISABLED_COMPONENTS = {
    "domain_pair",
    "functional_complementarity",
    "fusion",
    "known_interactions",
}
INVARIANT_TSV_FIELDS = (
    "query_id",
    "candidate_id",
    "candidate_description",
    "candidate_disposition",
    "disposition_reasons",
    "sequence_length",
    "gene_id",
    "locus_tag",
    "old_locus_tag",
    "contig",
    "strand",
    "has_coordinate",
    "has_annotation",
    "same_contig_as_query",
    "is_duplicate_sequence",
    "duplicate_group_id",
    "is_fragment_candidate",
    "fragment_reasons",
    "is_hypothetical",
    "identifier_match_status",
    "same_contig",
    "query_start",
    "query_end",
    "query_strand",
    "candidate_start",
    "candidate_end",
    "candidate_strand",
    "strand_relationship",
    "relative_position",
    "coordinate_position",
    "distance_bp",
    "overlap_bp",
    "intervening_gene_count",
    "intervening_feature_count",
    "feature_index_delta",
    "within_neighborhood_window",
    "context_completeness",
    "gene_context_status",
    "edge_to_edge_distance_bp",
    "within_neighborhood_gene_count",
    "operon_status",
    "operon_proxy_status",
    "operon_same_contig",
    "operon_same_strand",
    "operon_is_adjacent",
    "operon_intergenic_distance_bp",
    "operon_overlap_bp",
    "operon_intervening_gene_count",
    "operon_transcriptional_order",
    "operon_passes_distance_threshold",
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--baseline", type=Path, required=True)
    parser.add_argument("--single-root", type=Path, required=True)
    parser.add_argument("--scoring", type=Path, required=True)
    parser.add_argument("--repeat", type=Path, required=True)
    parser.add_argument("--shuffled", type=Path, required=True)
    parser.add_argument("--final", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    return parser.parse_args()


def read_tsv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def read_bundles(directory: Path) -> list[dict[str, Any]]:
    with (directory / "candidate_evidence_bundle.jsonl").open(encoding="utf-8") as handle:
        return [json.loads(line) for line in handle]


def close(left: float | None, right: float | None, tolerance: float = 1e-10) -> bool:
    if left is None or right is None:
        return left is right
    return math.isclose(left, right, rel_tol=0.0, abs_tol=tolerance)


def first(records: list[dict[str, Any]]) -> dict[str, Any] | None:
    return records[0] if records else None


def status(records: list[dict[str, Any]]) -> str:
    if not records:
        return "not_run"
    statuses = [str(record["status"]) for record in records]
    return "available" if "available" in statuses else statuses[0]


def raw_components(bundle: dict[str, Any]) -> dict[str, tuple[str, Any, float | None]]:
    values: dict[str, tuple[str, Any, float | None]] = {}
    context = first(bundle["genome_context"])
    context_status = str(context["status"]) if context else "not_run"
    context_raw: Any = context.get("distance_bp") if context else None
    context_value: float | None = None
    if context and context_status == "available":
        if context.get("coordinate_position") == "same_feature":
            context_raw, context_value = "same_feature", 0.0
        elif (context.get("overlap_bp") or 0) > 0:
            context_raw, context_value = f"overlap:{context['overlap_bp']}", 0.0
        elif context.get("same_contig") is True and context.get("within_neighborhood_window"):
            context_value = 0.8 if context.get("intervening_gene_count") == 0 else 0.5
        elif context.get("same_contig") is True:
            context_value = 0.1
    values["genome_context"] = (context_status, context_raw, context_value)

    operon = first(bundle["operon"])
    operon_status = str(operon["status"]) if operon else "not_run"
    proxy = operon.get("proxy_status") if operon else None
    values["operon_proxy"] = (
        operon_status,
        proxy,
        {"supported": 1.0, "partial_support": 0.3, "not_supported": 0.0}.get(str(proxy))
        if proxy is not None
        else None,
    )

    localization = first(bundle["localization"])
    localization_status = str(localization["status"]) if localization else "not_run"
    compatibility = localization.get("compatibility") if localization else None
    localization_value = None
    if localization_status == "available":
        localization_value = (
            1.0 if compatibility is True else -0.25 if compatibility is False else 0.0
        )
    values["localization"] = (localization_status, compatibility, localization_value)

    orthology = bundle["orthology"]
    orthology_status = status(orthology)
    supported = any(record.get("pair_supported") is True for record in orthology)
    orthology_value = None
    if orthology_status == "available":
        orthology_value = 1.0 if supported else 0.0
        if supported and any(record.get("paralog_ambiguity") for record in orthology):
            orthology_value *= 0.9
    values["orthology"] = (
        orthology_status,
        supported if orthology else None,
        orthology_value,
    )

    profile = first(bundle["phylogenetic_profile"])
    profile_status = str(profile["status"]) if profile else "not_run"
    similarity = profile.get("profile_similarity") if profile else None
    profile_value = None
    if (
        profile
        and profile_status == "available"
        and "insufficient_informative_species" not in profile.get("conflicting_terms", [])
        and similarity is not None
        and profile.get("shared_presence_count") is not None
    ):
        shared_presence = profile["shared_presence_count"]
        concordant = shared_presence + (profile.get("shared_absence_count") or 0)
        profile_value = similarity * (shared_presence / concordant if concordant else 0.0)
        informative = profile.get("informative_species_count") or 0
        discordant = profile.get("discordant_count") or 0
        if informative >= 3 and discordant / informative >= 0.8:
            profile_value = -0.25
    values["phylogenetic_profile"] = (profile_status, similarity, profile_value)

    for name in DISABLED_COMPONENTS:
        values[name] = ("not_run", None, None)
    if bundle["candidate"]["identifier_match_status"] == "ambiguous_match":
        values = {
            name: (component_status, raw, value * 0.9 if value is not None and value > 0 else value)
            for name, (component_status, raw, value) in values.items()
        }
    return values


def independently_score(bundle: dict[str, Any]) -> dict[str, Any]:
    raw = raw_components(bundle)
    components: dict[str, dict[str, Any]] = {}
    by_category: dict[str, list[str]] = defaultdict(list)
    for name in COMPONENT_ORDER:
        component_status, raw_value, value = raw[name]
        applied = component_status == "available" and value is not None and WEIGHTS[name] != 0
        direction = (
            "unknown"
            if value is None
            else "positive"
            if value > 0
            else "negative"
            if value < 0
            else "neutral"
        )
        components[name] = {
            "status": component_status,
            "raw": raw_value,
            "value": value,
            "applied": applied,
            "direction": direction,
            "configured_weight": WEIGHTS[name],
            "category": CATEGORIES[name],
            "effective_weight": 0.0,
            "contribution": None,
        }
        if applied:
            by_category[CATEGORIES[name]].append(name)
    for category, names in by_category.items():
        configured_total = sum(WEIGHTS[name] for name in names)
        scale = min(1.0, CAPS[category] / configured_total)
        for name in names:
            effective = round(WEIGHTS[name] * scale, PRECISION)
            components[name]["effective_weight"] = effective
            components[name]["contribution"] = round(
                components[name]["value"] * effective, PRECISION
            )
    category_scores: dict[str, dict[str, float]] = {}
    for category in sorted(CAPS):
        names = by_category.get(category, [])
        available = round(sum(components[name]["effective_weight"] for name in names), PRECISION)
        subtotal = round(sum(components[name]["contribution"] or 0.0 for name in names), PRECISION)
        category_scores[category] = {
            "available_weight": available,
            "raw_weighted_sum": subtotal,
            "normalized_score": round(subtotal / available, PRECISION) if available else 0.0,
        }
    applied_components = [component for component in components.values() if component["applied"]]
    available = round(
        sum(component["effective_weight"] for component in applied_components), PRECISION
    )
    raw_sum = round(
        sum(component["contribution"] or 0.0 for component in applied_components), PRECISION
    )
    normalized = round(max(0.0, min(1.0, raw_sum / available)), PRECISION) if available else None
    provisional = round(normalized * OUTPUT_SCALE, PRECISION) if normalized is not None else None
    category_count = len(by_category)
    sufficient = available >= MINIMUM_WEIGHT and category_count >= MINIMUM_CATEGORIES
    return {
        "components": components,
        "category_scores": category_scores,
        "available_weight": available,
        "raw_weighted_sum": raw_sum,
        "normalized_score": normalized,
        "provisional_score": provisional,
        "output_score": provisional if sufficient else None,
        "category_count": category_count,
        "component_count": len(applied_components),
        "positive_count": sum(
            component["direction"] == "positive" for component in applied_components
        ),
        "neutral_count": sum(
            component["direction"] == "neutral" for component in applied_components
        ),
        "negative_count": sum(
            component["direction"] == "negative" for component in applied_components
        ),
        "sufficient": sufficient,
    }


def audit_recalculation(bundles: list[dict[str, Any]], output: Path) -> dict[str, int]:
    failures: list[dict[str, Any]] = []
    independent: dict[str, dict[str, Any]] = {}
    for bundle in bundles:
        candidate_id = bundle["candidate_id"]
        expected = independently_score(bundle)
        independent[candidate_id] = expected
        actual = bundle["integrated_scoring"][0]
        actual_components = {item["component_name"]: item for item in actual["component_scores"]}
        for name, component in expected["components"].items():
            observed = actual_components[name]
            checks = {
                "status": observed["evidence_status"] == component["status"],
                "raw": observed["raw_value"] == component["raw"],
                "normalized": close(observed["normalized_value"], component["value"]),
                "applied": observed["applied"] == component["applied"],
                "weight": close(observed["configured_weight"], component["configured_weight"]),
                "effective_weight": close(
                    observed["effective_weight"], component["effective_weight"]
                ),
                "contribution": close(observed["weighted_contribution"], component["contribution"]),
                "category": observed["category_name"] == component["category"],
            }
            if not all(checks.values()):
                failures.append({"candidate_id": candidate_id, "component": name, "checks": checks})
        actual_categories = {item["category_name"]: item for item in actual["category_scores"]}
        for name, category in expected["category_scores"].items():
            observed = actual_categories[name]
            if not all(
                (
                    close(observed["available_weight"], category["available_weight"]),
                    close(observed["raw_weighted_sum"], category["raw_weighted_sum"]),
                    close(observed["normalized_score"], category["normalized_score"]),
                    close(observed["configured_cap"], CAPS[name]),
                )
            ):
                failures.append({"candidate_id": candidate_id, "category": name})
        scalar_checks = {
            "available_weight": close(actual["available_weight"], expected["available_weight"]),
            "raw_weighted_sum": close(actual["raw_weighted_sum"], expected["raw_weighted_sum"]),
            "normalized_score": close(actual["normalized_score"], expected["normalized_score"]),
            "provisional_score": close(actual["provisional_score"], expected["provisional_score"]),
            "output_score": close(actual["output_score"], expected["output_score"]),
            "category_count": actual["evidence_category_count"] == expected["category_count"],
            "component_count": actual["evidence_component_count"] == expected["component_count"],
            "positive_count": actual["positive_component_count"] == expected["positive_count"],
            "neutral_count": actual["neutral_component_count"] == expected["neutral_count"],
            "negative_count": actual["negative_component_count"] == expected["negative_count"],
            "sufficient": actual["sufficient_evidence"] == expected["sufficient"],
        }
        if not all(scalar_checks.values()):
            failures.append({"candidate_id": candidate_id, "score": scalar_checks})

    eligible = [
        (candidate_id, score)
        for candidate_id, score in independent.items()
        if next(bundle for bundle in bundles if bundle["candidate_id"] == candidate_id)[
            "candidate_disposition"
        ]
        != "excluded"
        and score["output_score"] is not None
    ]
    eligible.sort(
        key=lambda item: (
            -round(item[1]["output_score"], TIE_PRECISION),
            -item[1]["category_count"],
            -item[1]["available_weight"],
            item[0],
        )
    )
    dense_rank = 0
    previous = None
    expected_ranks: dict[str, int] = {}
    for candidate_id, score in eligible:
        tie_value = round(score["output_score"], TIE_PRECISION)
        if previous is None or tie_value != previous:
            dense_rank += 1
            previous = tie_value
        expected_ranks[candidate_id] = dense_rank
    for bundle in bundles:
        actual_rank = bundle["integrated_scoring"][0]["rank"]
        if actual_rank != expected_ranks.get(bundle["candidate_id"]):
            failures.append(
                {
                    "candidate_id": bundle["candidate_id"],
                    "rank": {
                        "expected": expected_ranks.get(bundle["candidate_id"]),
                        "actual": actual_rank,
                    },
                }
            )
    (output / "recalculation_failures.json").write_text(
        json.dumps(failures, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )
    return {
        "candidate_count": len(bundles),
        "component_comparisons": len(bundles) * len(COMPONENT_ORDER),
        "failure_count": len(failures),
        "ranked_count": len(expected_ranks),
    }


def canonical_scoring_rows(directory: Path) -> list[tuple[str, ...]]:
    fields = (
        "candidate_id",
        "candidate_disposition",
        "integrated_score",
        "provisional_score",
        "normalized_score",
        "rank",
        "tied_rank",
        "sufficient_evidence",
        "available_weight",
        "evidence_category_count",
        "evidence_component_count",
        "positive_component_count",
        "neutral_component_count",
        "negative_component_count",
        "scoring_rule_version",
        *[f"score_component_{name}" for name in COMPONENT_ORDER],
    )
    return [
        tuple(row[field] for field in fields) for row in read_tsv(directory / "candidate_table.tsv")
    ]


def excel_score_rank(directory: Path) -> list[tuple[Any, ...]]:
    from openpyxl import load_workbook

    workbook = load_workbook(
        directory / "ProteinInteractionHunter.xlsx", read_only=True, data_only=True
    )
    sheet = workbook["Integrated_Scoring"]
    rows = sheet.iter_rows(values_only=True)
    header = tuple(next(rows))
    indexes = [header.index(name) for name in ("Candidate_ID", "Integrated_Score", "Rank")]
    result = [tuple(row[index] for index in indexes) for row in rows]
    workbook.close()
    return result


def determinism(scoring: Path, repeat: Path, shuffled: Path) -> dict[str, Any]:
    reference = canonical_scoring_rows(scoring)
    reference_excel = excel_score_rank(scoring)
    return {
        "repeat_tsv_identical": reference == canonical_scoring_rows(repeat),
        "shuffled_tsv_identical": reference == canonical_scoring_rows(shuffled),
        "repeat_excel_identical": reference_excel == excel_score_rank(repeat),
        "shuffled_excel_identical": reference_excel == excel_score_rank(shuffled),
        "candidate_order_identical": [row[0] for row in reference]
        == [row[0] for row in canonical_scoring_rows(shuffled)],
    }


def count_values(rows: list[dict[str, str]], field: str) -> dict[str, int]:
    return dict(sorted(Counter(row[field] or "<empty>" for row in rows).items()))


def quantile(values: list[float], fraction: float) -> float | None:
    if not values:
        return None
    position = (len(values) - 1) * fraction
    lower = math.floor(position)
    upper = math.ceil(position)
    if lower == upper:
        return values[lower]
    return values[lower] + (values[upper] - values[lower]) * (position - lower)


def score_summary(rows: list[dict[str, str]], bundles: list[dict[str, Any]]) -> dict[str, Any]:
    scored = [float(row["integrated_score"]) for row in rows if row["integrated_score"]]
    positive = [value for value in scored if value > 0]
    rounded = Counter(round(value, TIE_PRECISION) for value in scored)
    tie_groups = [count for count in rounded.values() if count > 1]
    penalties = Counter(
        penalty["penalty_name"]
        for bundle in bundles
        for score in bundle["integrated_scoring"]
        for penalty in score["applied_penalties"]
    )
    sorted_scores = sorted(scored)
    return {
        "score_generated": len(scored),
        "score_none": len(rows) - len(scored),
        "score_zero": sum(value == 0 for value in scored),
        "score_positive": len(positive),
        "minimum": min(scored) if scored else None,
        "minimum_positive": min(positive) if positive else None,
        "maximum": max(scored) if scored else None,
        "median": statistics.median(scored) if scored else None,
        "mean": statistics.fmean(scored) if scored else None,
        "quantiles": {
            str(fraction): quantile(sorted_scores, fraction)
            for fraction in (0.0, 0.25, 0.5, 0.75, 0.9, 0.95, 0.99, 1.0)
        },
        "unique_score_count": len(rounded),
        "tie_group_count": len(tie_groups),
        "largest_tie_group": max(tie_groups, default=1),
        "available_weight": count_values(rows, "available_weight"),
        "category_count": count_values(rows, "evidence_category_count"),
        "component_count": count_values(rows, "evidence_component_count"),
        "negative_count": count_values(rows, "negative_component_count"),
        "penalties": dict(sorted(penalties.items())),
        "minimum_category_failure": sum(
            row["evidence_category_count"]
            and int(row["evidence_category_count"]) < MINIMUM_CATEGORIES
            for row in rows
        ),
        "minimum_weight_failure": sum(
            row["available_weight"] and float(row["available_weight"]) < MINIMUM_WEIGHT
            for row in rows
        ),
    }


def baseline_audit(baseline: Path, final: Path) -> dict[str, Any]:
    before = {row["candidate_id"]: row for row in read_tsv(baseline / "candidate_table.tsv")}
    after = {row["candidate_id"]: row for row in read_tsv(final / "candidate_table.tsv")}
    differences: list[dict[str, str]] = []
    for candidate_id in sorted(set(before) & set(after)):
        for field in INVARIANT_TSV_FIELDS:
            if before[candidate_id][field] != after[candidate_id][field]:
                differences.append(
                    {
                        "candidate_id": candidate_id,
                        "field": field,
                        "before": before[candidate_id][field],
                        "after": after[candidate_id][field],
                    }
                )
    baseline_bundles = {item["candidate_id"]: item for item in read_bundles(baseline)}
    final_bundles = {item["candidate_id"]: item for item in read_bundles(final)}
    raw_json_differences = 0
    for candidate_id in sorted(set(baseline_bundles) & set(final_bundles)):
        left, right = baseline_bundles[candidate_id], final_bundles[candidate_id]
        for field in ("candidate", "genome_context", "operon", "candidate_disposition"):
            if left[field] != right[field]:
                raw_json_differences += 1
    return {
        "baseline_count": len(before),
        "final_count": len(after),
        "candidate_id_sets_identical": set(before) == set(after),
        "invariant_tsv_difference_count": len(differences),
        "invariant_json_difference_count": raw_json_differences,
        "unexpected_examples": differences[:20],
    }


def write_tier_audit(rows: list[dict[str, str]], path: Path) -> None:
    fields = (
        "protein_id",
        "score",
        "rank",
        "raw_eligible_tier",
        "final_tier",
        "category_count",
        "component_count",
        "available_weight",
        "negative_component_count",
        "explicit_conflict_cap",
        "predicted_only_cap",
        "functional_association_only_cap",
        "satisfied_requirements",
        "failed_requirements",
        "final_reason",
        "rule_version",
    )
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, delimiter="\t", fieldnames=fields)
        writer.writeheader()
        for row in rows:
            caps = set(filter(None, row["evidence_tier_applied_caps"].split("|")))
            writer.writerow(
                {
                    "protein_id": row["candidate_id"],
                    "score": row["integrated_score"],
                    "rank": row["rank"],
                    "raw_eligible_tier": row["evidence_tier_base"],
                    "final_tier": row["evidence_tier"],
                    "category_count": row["evidence_tier_category_count"],
                    "component_count": row["evidence_tier_component_count"],
                    "available_weight": row["evidence_tier_available_weight"],
                    "negative_component_count": row["evidence_tier_negative_component_count"],
                    "explicit_conflict_cap": next(
                        (value for value in caps if value.startswith("explicit_conflict:")), ""
                    ),
                    "predicted_only_cap": next(
                        (value for value in caps if value.startswith("predicted_only:")), ""
                    ),
                    "functional_association_only_cap": next(
                        (
                            value
                            for value in caps
                            if value.startswith("functional_association_only:")
                        ),
                        "",
                    ),
                    "satisfied_requirements": row["evidence_tier_satisfied_requirements"],
                    "failed_requirements": row["evidence_tier_failed_requirements"],
                    "final_reason": (
                        row["evidence_tier_failed_requirements"]
                        or row["evidence_tier_applied_caps"]
                        or f"base_tier:{row['evidence_tier_base']}"
                    ),
                    "rule_version": row["evidence_tier_rule_version"],
                }
            )


def write_review(rows: list[dict[str, str]], path: Path, maximum_rank: int) -> int:
    ranked = [row for row in rows if row["rank"] and int(row["rank"]) <= maximum_rank]
    fields = (
        "rank",
        "candidate_id",
        "locus_tag",
        "old_locus_tag",
        "candidate_description",
        "sequence_length",
        "candidate_disposition",
        "is_fragment_candidate",
        "identifier_match_status",
        "distance_bp",
        "candidate_strand",
        "intervening_gene_count",
        "gene_context_status",
        "score_component_genome_context",
        "operon_proxy_status",
        "localization_compartment",
        "localization_status",
        "orthology_status",
        "score_component_orthology",
        "phylogenetic_profile_status",
        "phylogenetic_profile_similarity",
        "phylogenetic_profile_shared_presence",
        "phylogenetic_profile_informative_species",
        "functional_status",
        "domain_status",
        "evidence_category_count",
        "positive_component_count",
        "negative_component_count",
        "available_weight",
        "integrated_score",
        "evidence_tier",
        "evidence_tier_failed_requirements",
        "warnings",
    )
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, delimiter="\t", fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(ranked)
    return len(ranked)


def single_engine_summary(root: Path) -> dict[str, Any]:
    result: dict[str, Any] = {}
    mapping = {
        "gene_context": ("gene_context_status", "operon_proxy_status"),
        "localization": ("localization_status", "localization_compatibility"),
        "orthology": ("orthology_status", "orthology_pair_supported"),
        "profile": ("phylogenetic_profile_status", "phylogenetic_profile_pair_supported"),
    }
    for run, fields in mapping.items():
        rows = read_tsv(root / run / "candidate_table.tsv")
        result[run] = {
            "candidate_count": len(rows),
            "disposition": count_values(rows, "candidate_disposition"),
            **{field: count_values(rows, field) for field in fields},
            "scoring_not_generated": all(not row["scoring_status"] for row in rows),
            "tier_not_generated": all(not row["evidence_tier_status"] for row in rows),
        }
    return result


def main() -> int:
    args = parse_args()
    args.output.mkdir(parents=True, exist_ok=True)
    final_rows = read_tsv(args.final / "candidate_table.tsv")
    final_bundles = read_bundles(args.final)
    recalculation = audit_recalculation(final_bundles, args.output)
    rank_determinism = determinism(args.scoring, args.repeat, args.shuffled)
    baseline = baseline_audit(args.baseline, args.final)
    top25_count = write_review(final_rows, args.output / "top25_with_boundary_ties.tsv", 25)
    top50_count = write_review(final_rows, args.output / "top50_with_boundary_ties.tsv", 50)
    write_tier_audit(final_rows, args.output / "tier_decision_audit.tsv")
    target = next(row for row in final_rows if row["candidate_id"] == "WP_011024007.1")
    summary = {
        "single_engines": single_engine_summary(args.single_root),
        "recalculation": recalculation,
        "determinism": rank_determinism,
        "baseline_ab": baseline,
        "score": score_summary(final_rows, final_bundles),
        "tier_distribution": count_values(final_rows, "evidence_tier"),
        "tier_base_distribution": count_values(final_rows, "evidence_tier_base"),
        "predicted_only": count_values(final_rows, "evidence_tier_predicted_only"),
        "explicit_conflict": count_values(final_rows, "evidence_tier_explicit_conflict"),
        "applied_tier_caps": count_values(final_rows, "evidence_tier_applied_caps"),
        "disposition": count_values(final_rows, "candidate_disposition"),
        "localization": {
            "status": count_values(final_rows, "localization_status"),
            "compatibility": count_values(final_rows, "localization_compatibility"),
            "compartment": count_values(final_rows, "localization_compartment"),
        },
        "orthology": {
            "status": count_values(final_rows, "orthology_status"),
            "supported": count_values(final_rows, "orthology_pair_supported"),
        },
        "profile": {
            "status": count_values(final_rows, "phylogenetic_profile_status"),
            "supported": count_values(final_rows, "phylogenetic_profile_pair_supported"),
        },
        "top_review_counts": {"top25_with_ties": top25_count, "top50_with_ties": top50_count},
        "wp_011024007_1": {
            field: target[field]
            for field in (
                "rank",
                "integrated_score",
                "evidence_tier",
                "distance_bp",
                "operon_proxy_status",
                "localization_compartment",
                "orthology_pair_supported",
                "phylogenetic_profile_similarity",
            )
        },
    }
    gates = {
        "recalculation": recalculation["failure_count"] == 0,
        "repeat_and_shuffle": all(rank_determinism.values()),
        "baseline": (
            baseline["candidate_id_sets_identical"]
            and baseline["invariant_tsv_difference_count"] == 0
            and baseline["invariant_json_difference_count"] == 0
        ),
        "candidate_count": len(final_rows) == 4627,
        "excluded_not_ranked": all(
            not row["rank"] for row in final_rows if row["candidate_disposition"] == "excluded"
        ),
        "score_none_not_ranked": all(
            not row["rank"] for row in final_rows if not row["integrated_score"]
        ),
        "disabled_components_not_run": all(
            all(
                next(
                    component
                    for component in bundle["integrated_scoring"][0]["component_scores"]
                    if component["component_name"] == name
                )["evidence_status"]
                == "not_run"
                for name in DISABLED_COMPONENTS
            )
            for bundle in final_bundles
        ),
    }
    summary["gates"] = gates
    (args.output / "summary.json").write_text(
        json.dumps(summary, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )
    print(json.dumps({"gates": gates, "summary": summary}, indent=2, sort_keys=True))
    return 0 if all(gates.values()) else 1


if __name__ == "__main__":
    raise SystemExit(main())
