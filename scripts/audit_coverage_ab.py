#!/usr/bin/env python3
"""Compare two coverage-only outputs while excluding run metadata."""

from __future__ import annotations

import argparse
import csv
import json
from collections.abc import Sequence
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from protein_interaction_hunter.exceptions import InputValidationError


def _candidate_rows(path: Path) -> tuple[tuple[str, ...], list[dict[str, str]]]:
    with path.open(encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        header = tuple(reader.fieldnames or ())
        rows = [dict(row) for row in reader]
    for row in rows:
        row["run_id"] = ""
    return header, rows


def _jsonl(path: Path) -> list[dict[str, Any]]:
    rows = [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line]
    for row in rows:
        row["run_id"] = ""
    return rows


def _excel_shape(path: Path) -> tuple[tuple[str, int, int], ...]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        return tuple(
            (worksheet.title, worksheet.max_row, worksheet.max_column)
            for worksheet in workbook.worksheets
        )
    finally:
        workbook.close()


def _changed_manifest_keys(left: object, right: object, prefix: str = "") -> list[str]:
    if isinstance(left, dict) and isinstance(right, dict):
        keys = sorted(set(left) | set(right))
        result: list[str] = []
        for key in keys:
            child = f"{prefix}.{key}" if prefix else key
            result.extend(_changed_manifest_keys(left.get(key), right.get(key), child))
        return result
    return [] if left == right else [prefix]


def audit_ab(
    baseline: Path,
    comparison: Path,
) -> tuple[tuple[str, str, str], ...]:
    baseline_header, baseline_rows = _candidate_rows(baseline / "candidate_table.tsv")
    comparison_header, comparison_rows = _candidate_rows(comparison / "candidate_table.tsv")
    candidate_equal = baseline_header == comparison_header and baseline_rows == comparison_rows
    baseline_jsonl = _jsonl(baseline / "candidate_evidence_bundle.jsonl")
    comparison_jsonl = _jsonl(comparison / "candidate_evidence_bundle.jsonl")
    jsonl_equal = baseline_jsonl == comparison_jsonl
    warnings_equal = (baseline / "warning_summary.tsv").read_bytes() == (
        comparison / "warning_summary.tsv"
    ).read_bytes()
    excel_equal = _excel_shape(baseline / "ProteinInteractionHunter.xlsx") == (
        _excel_shape(comparison / "ProteinInteractionHunter.xlsx")
    )

    status_ok = all(
        row["functional_status"] in {"", "not_run"}
        and row["scoring_status"] in {"", "not_run"}
        and row["evidence_tier_status"] in {"", "not_run"}
        and not row["integrated_score"]
        and not row["rank"]
        and not row["evidence_tier"]
        for row in comparison_rows
    )
    bundle_status_ok = all(
        row["engine_statuses"]["functional_complementarity"] == "not_run"
        and row["engine_statuses"]["scoring"] == "not_run"
        and row["engine_statuses"]["evidence_tiers"] == "not_run"
        and not row["functional"]
        and row["score"]["total_ranking_score"] is None
        and row["evidence_tier"] is None
        for row in comparison_jsonl
    )
    baseline_manifest = json.loads((baseline / "run_manifest.json").read_text(encoding="utf-8"))
    comparison_manifest = json.loads((comparison / "run_manifest.json").read_text(encoding="utf-8"))
    changed_keys = _changed_manifest_keys(baseline_manifest, comparison_manifest)
    results = (
        ("candidate_tsv_excluding_run_id", str(candidate_equal), ""),
        ("jsonl_excluding_run_id", str(jsonl_equal), ""),
        ("warning_summary", str(warnings_equal), ""),
        ("excel_sheet_names_and_dimensions", str(excel_equal), ""),
        ("tsv_functional_score_rank_tier_not_run", str(status_ok), ""),
        ("jsonl_functional_score_rank_tier_not_run", str(bundle_status_ok), ""),
        (
            "candidate_pair_count",
            str(len(comparison_rows) == len(baseline_rows)),
            str(len(comparison_rows)),
        ),
        (
            "manifest_changed_keys",
            "True",
            "|".join(changed_keys),
        ),
    )
    if not all(
        value == "True" for check, value, _details in results if check != "manifest_changed_keys"
    ):
        raise InputValidationError("Coverage-only A/B comparison failed")
    return results


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--baseline", required=True, type=Path)
    parser.add_argument("--comparison", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    args = _parser().parse_args(argv)
    rows = audit_ab(args.baseline, args.comparison)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    with args.output.open("w", encoding="utf-8", newline="\n") as handle:
        writer = csv.writer(handle, delimiter="\t", lineterminator="\n")
        writer.writerow(("check", "passed", "details"))
        writer.writerows(rows)
    print("Coverage-only A/B checks passed")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
