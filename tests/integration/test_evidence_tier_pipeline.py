import csv
from copy import deepcopy
from pathlib import Path

import yaml
from openpyxl import load_workbook

from protein_interaction_hunter.application.evidence_tiers import (
    EVIDENCE_TIER_RULE_VERSION,
    HIGH_SPECIFICITY_DEFINITION_VERSION,
)
from protein_interaction_hunter.application.pipeline import InteractionCandidatePipeline
from protein_interaction_hunter.application.scoring import SCORING_RULE_VERSION
from protein_interaction_hunter.models.enums import EvidenceStatus, EvidenceTier
from protein_interaction_hunter.models.run import RunManifest


def config_for_run(source: Path, tmp_path: Path, *, tiers: bool) -> Path:
    data = yaml.safe_load(source.read_text(encoding="utf-8"))
    fixture = source.parent
    for key in ("proteome_fasta", "genome_gff", "annotation_table"):
        value = data["input"].get(key)
        if value:
            data["input"][key] = str((fixture / value).resolve())
    for section in ("orthology", "phylogenetic_profile", "fusion", "domains", "known_interactions"):
        value = data[section].get("local_table")
        if value:
            data[section]["local_table"] = str((fixture / value).resolve())
    for section in ("domains", "functional_complementarity"):
        value = data[section].get("rules_path")
        if value:
            data[section]["rules_path"] = str((fixture / value).resolve())
    data["localization"]["enabled"] = True
    data["scoring"]["enabled"] = True
    data["evidence_tiers"]["enabled"] = tiers
    data["output"]["directory"] = str(tmp_path / ("enabled" if tiers else "disabled"))
    path = tmp_path / ("enabled.yaml" if tiers else "disabled.yaml")
    path.write_text(yaml.safe_dump(data, sort_keys=False), encoding="utf-8")
    return path


def test_evidence_tier_pipeline_enabled_disabled_ab(
    valid_config_path: Path, tmp_path: Path
) -> None:
    disabled = InteractionCandidatePipeline().run(
        config_for_run(valid_config_path, tmp_path, tiers=False)
    )
    enabled = InteractionCandidatePipeline().run(
        config_for_run(valid_config_path, tmp_path, tiers=True)
    )

    assert [item.candidate_id for item in disabled.bundles] == [
        item.candidate_id for item in enabled.bundles
    ]
    for before, after in zip(disabled.bundles, enabled.bundles, strict=True):
        scoring_before = deepcopy(before.integrated_scoring)
        raw_before = {
            name: deepcopy(getattr(before, name))
            for name in (
                "genome_context",
                "operon",
                "domains",
                "functional",
                "localization",
                "orthology",
                "phylogenetic_profile",
                "fusion",
                "known_interactions",
            )
        }
        assert before.evidence_tier is None
        assert before.evidence_tiers == []
        assert before.engine_statuses["evidence_tiers"] is EvidenceStatus.NOT_RUN
        assert len(after.evidence_tiers) == 1
        assert after.evidence_tier is after.evidence_tiers[0].assigned_tier
        assert after.engine_statuses["evidence_tiers"] is after.evidence_tiers[0].status
        assert before.candidate_disposition == after.candidate_disposition
        assert before.predicted_relationship_type == after.predicted_relationship_type
        assert before.integrated_scoring == after.integrated_scoring == scoring_before
        assert before.score == after.score
        for name, value in raw_before.items():
            assert getattr(after, name) == value

    enabled_by_id = {item.candidate_id: item for item in enabled.bundles}
    assert enabled_by_id["QUERY_001"].evidence_tier is EvidenceTier.UNCLASSIFIED
    assert enabled_by_id["QUERY_001"].evidence_tiers[0].tier_eligible is False
    assert any(
        item.evidence_tiers[0].high_specificity_components
        for item in enabled.bundles
        if item.evidence_tiers[0].tier_eligible
    )

    with disabled.candidate_table_path.open(encoding="utf-8", newline="") as handle:
        disabled_rows = list(csv.DictReader(handle, delimiter="\t"))
    with enabled.candidate_table_path.open(encoding="utf-8", newline="") as handle:
        enabled_rows = list(csv.DictReader(handle, delimiter="\t"))
    assert [row["candidate_id"] for row in disabled_rows] == [
        row["candidate_id"] for row in enabled_rows
    ]
    assert all(row["evidence_tier"] == "" for row in disabled_rows)
    assert all(row["evidence_tier"] for row in enabled_rows)
    for row_before, row_after in zip(disabled_rows, enabled_rows, strict=True):
        for field in (
            "candidate_disposition",
            "integrated_score",
            "rank",
            "scoring_status",
            "scoring_rule_version",
            "scoring_support_terms",
            "scoring_conflicting_terms",
            "scoring_warnings",
        ):
            assert row_before[field] == row_after[field]

    assert enabled.excel_path is not None
    workbook = load_workbook(enabled.excel_path, read_only=True, data_only=True)
    assert workbook["Evidence_Tiers"].max_row == len(enabled.bundles) + 1
    assert workbook["Tier_Summary"].max_row == 2
    ranking_headers = [
        cell.value for cell in next(workbook["Candidate_Ranking"].iter_rows(max_row=1))
    ]
    assert "Evidence_Tier" in ranking_headers
    workbook.close()

    assert disabled.excel_path is not None
    workbook = load_workbook(disabled.excel_path, read_only=True, data_only=True)
    assert workbook["Evidence_Tiers"].max_row == 1
    assert workbook["Tier_Summary"].max_row == 1
    workbook.close()

    enabled_manifest = RunManifest.model_validate_json(
        enabled.manifest_path.read_text(encoding="utf-8")
    )
    assert enabled_manifest.evidence_tier_rule_version == EVIDENCE_TIER_RULE_VERSION
    assert enabled_manifest.evidence_tier_config_snapshot["enabled"] is True
    assert enabled_manifest.evidence_tier_thresholds["tier_1"]["minimum_score"] == 75.0
    assert enabled_manifest.evidence_tier_caps["predicted_only"] == "tier_3"
    assert (
        enabled_manifest.evidence_tier_high_specificity_definition_version
        == HIGH_SPECIFICITY_DEFINITION_VERSION
    )
    assert enabled_manifest.evidence_tier_scoring_rule_version_dependency == SCORING_RULE_VERSION
    assert "evidence_tiers_not_run" not in enabled_manifest.incomplete_evidence_flags

    disabled_manifest = RunManifest.model_validate_json(
        disabled.manifest_path.read_text(encoding="utf-8")
    )
    assert disabled_manifest.evidence_tier_rule_version is None
    assert disabled_manifest.evidence_tier_config_snapshot == {}
    assert "evidence_tiers_not_run" in disabled_manifest.incomplete_evidence_flags


def test_semantic_caps_are_visible_in_derived_outputs(
    valid_config_path: Path, tmp_path: Path
) -> None:
    config_path = config_for_run(valid_config_path, tmp_path, tiers=True)
    data = yaml.safe_load(config_path.read_text(encoding="utf-8"))
    for name in ("tier_2", "tier_3", "tier_4"):
        threshold = data["evidence_tiers"][name]
        threshold.update(
            {
                "minimum_score": 0.0,
                "minimum_categories": 2,
                "minimum_components": 2,
                "minimum_available_weight": 1.0,
                "require_high_specificity_evidence": False,
                "minimum_high_specificity_components": 0,
                "maximum_negative_components": 99,
            }
        )
    config_path.write_text(yaml.safe_dump(data, sort_keys=False), encoding="utf-8")
    result = InteractionCandidatePipeline().run(config_path)
    by_id = {item.candidate_id: item.evidence_tiers[0] for item in result.bundles}

    # MID has a functional-association record but no direct/physical/fusion support.
    assert by_id["MID_001"].base_tier is EvidenceTier.TIER_2
    assert by_id["MID_001"].assigned_tier is EvidenceTier.TIER_3
    assert by_id["MID_001"].applied_tier_caps == ["functional_association_only:tier_3"]
    # FRAG's known-interaction record is predicted-only.
    assert by_id["FRAG_001"].base_tier is EvidenceTier.TIER_2
    assert by_id["FRAG_001"].assigned_tier is EvidenceTier.TIER_3
    assert by_id["FRAG_001"].applied_tier_caps == ["predicted_only:tier_3"]

    with result.candidate_table_path.open(encoding="utf-8", newline="") as handle:
        rows = {row["candidate_id"]: row for row in csv.DictReader(handle, delimiter="\t")}
    assert rows["MID_001"]["evidence_tier_applied_caps"] == ("functional_association_only:tier_3")
    assert rows["FRAG_001"]["evidence_tier_applied_caps"] == "predicted_only:tier_3"

    assert result.excel_path is not None
    workbook = load_workbook(result.excel_path, read_only=True, data_only=True)
    sheet = workbook["Evidence_Tiers"]
    excel_rows = list(sheet.iter_rows(values_only=True))
    headers = list(excel_rows[0])
    candidate_index = headers.index("Candidate_ID")
    caps_index = headers.index("Applied_Tier_Caps")
    cap_by_id = {row[candidate_index]: row[caps_index] for row in excel_rows[1:]}
    assert cap_by_id["MID_001"] == "functional_association_only:tier_3"
    assert cap_by_id["FRAG_001"] == "predicted_only:tier_3"
    workbook.close()
