"""Canonical JSONL and derived TSV/Excel schema tests."""

import csv
import json
from pathlib import Path

from openpyxl import load_workbook

from protein_interaction_hunter.models import (
    CandidateDisposition,
    CandidateEvidenceBundle,
    EvidenceStatus,
    EvidenceTier,
    ManualStructurePriority,
    PredictedRelationshipType,
    StructurePredictionQueueEntry,
)
from protein_interaction_hunter.outputs.excel import EXCEL_SHEETS, ExcelSchemaWriter
from protein_interaction_hunter.outputs.jsonl import JsonlEvidenceBundleWriter
from protein_interaction_hunter.outputs.tsv import (
    STRUCTURE_QUEUE_COLUMNS,
    StructureQueueTsvWriter,
)


def bundle(candidate_id: str) -> CandidateEvidenceBundle:
    return CandidateEvidenceBundle(
        run_id="fixture-run",
        query_id="QUERY_001",
        candidate_id=candidate_id,
        candidate_disposition=CandidateDisposition.INCLUDED,
        predicted_relationship_type=PredictedRelationshipType.INSUFFICIENT_EVIDENCE,
        engine_statuses={
            "gene_context": EvidenceStatus.NOT_RUN,
            "orthology": EvidenceStatus.NOT_RUN,
        },
    )


def queue_entry() -> StructurePredictionQueueEntry:
    return StructurePredictionQueueEntry(
        rank=1,
        query_id="QUERY_001",
        candidate_id="NEAR_001",
        candidate_name="nearby accessory candidate",
        evidence_tier=EvidenceTier.TIER_C_EXPLORATORY,
        predicted_relationship_type=PredictedRelationshipType.ACCESSORY_FACTOR,
        manual_structure_priority=ManualStructurePriority.MEDIUM,
        suggested_stoichiometry="A1:B1 (hypothesis)",
        query_fasta_path=Path("queries/QUERY_001.fasta"),
        candidate_fasta_path=Path("candidates/NEAR_001.fasta"),
        pair_fasta_path=Path("pairs/QUERY_001__NEAR_001.fasta"),
        reason_for_structural_test="Manual hypothesis review only",
        primary_supporting_evidence="Synthetic fixture context",
    )


def test_expected_jsonl_line_is_independently_parseable(fixture_dir: Path) -> None:
    line = (
        fixture_dir
        / "expected"
        / "candidate_evidence_bundle.jsonl"
    ).read_text(encoding="utf-8").strip()
    parsed = CandidateEvidenceBundle.model_validate_json(line)
    assert parsed.engine_statuses["gene_context"] is EvidenceStatus.NOT_RUN


def test_jsonl_round_trip_is_deterministic(tmp_path: Path) -> None:
    path = tmp_path / "evidence.jsonl"
    writer = JsonlEvidenceBundleWriter()
    writer.write([bundle("Z_001"), bundle("A_001")], path)
    lines = path.read_text(encoding="utf-8").splitlines()
    assert json.loads(lines[0])["candidate_id"] == "A_001"
    assert writer.read(path) == [bundle("A_001"), bundle("Z_001")]
    assert all(len(line) > 0 for line in lines)


def test_structure_queue_tsv_header_and_expected_fixture(
    fixture_dir: Path, tmp_path: Path
) -> None:
    path = tmp_path / "queue.tsv"
    StructureQueueTsvWriter().write([queue_entry()], path)
    with path.open(encoding="utf-8", newline="") as handle:
        reader = csv.reader(handle, delimiter="\t")
        assert tuple(next(reader)) == STRUCTURE_QUEUE_COLUMNS
        assert next(reader)[2:4] == ["QUERY_001", "NEAR_001"]
    expected_header = (
        fixture_dir / "expected" / "structure_prediction_queue.tsv"
    ).read_text(encoding="utf-8").splitlines()[0]
    assert expected_header.split("\t") == list(STRUCTURE_QUEUE_COLUMNS)


def test_excel_schema_sheets_and_headers(tmp_path: Path) -> None:
    path = ExcelSchemaWriter().write(tmp_path / "schema.xlsx")
    workbook = load_workbook(path, read_only=True)
    assert workbook.sheetnames == list(EXCEL_SHEETS)
    for sheet_name, expected_headers in EXCEL_SHEETS.items():
        worksheet = workbook[sheet_name]
        headers = tuple(cell.value for cell in next(worksheet.iter_rows(max_row=1)))
        assert headers == expected_headers
    workbook.close()
