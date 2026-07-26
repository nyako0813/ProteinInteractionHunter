"""Candidate pipeline artifacts, provenance, and CLI end-to-end behavior."""

import csv
import json
from pathlib import Path

import pytest
import yaml
from openpyxl import load_workbook
from typer.testing import CliRunner

from protein_interaction_hunter.application.fusion import FUSION_ENGINE_VERSION
from protein_interaction_hunter.application.gene_context import GENE_CONTEXT_RULE_VERSION
from protein_interaction_hunter.application.identifiers import NORMALIZATION_RULE_VERSION
from protein_interaction_hunter.application.known_interactions import (
    KNOWN_INTERACTIONS_ENGINE_VERSION,
)
from protein_interaction_hunter.application.localization import (
    LOCALIZATION_ENGINE_VERSION,
)
from protein_interaction_hunter.application.orthology import (
    ORTHOLOGY_ENGINE_VERSION,
)
from protein_interaction_hunter.application.phylogenetic_profile import (
    PHYLOGENETIC_PROFILE_ENGINE_VERSION,
)
from protein_interaction_hunter.application.pipeline import InteractionCandidatePipeline
from protein_interaction_hunter.application.scoring import SCORING_RULE_VERSION
from protein_interaction_hunter.cli import app
from protein_interaction_hunter.exceptions import InputValidationError
from protein_interaction_hunter.models import CandidateEvidenceBundle, EvidenceStatus, RunManifest
from protein_interaction_hunter.outputs.candidates import CANDIDATE_COLUMNS
from protein_interaction_hunter.outputs.jsonl import JsonlEvidenceBundleWriter


def e2e_config(source: Path, tmp_path: Path) -> Path:
    data = yaml.safe_load(source.read_text(encoding="utf-8"))
    fixture_dir = source.parent

    data["input"]["proteome_fasta"] = str((fixture_dir / data["input"]["proteome_fasta"]).resolve())
    data["input"]["genome_gff"] = str((fixture_dir / data["input"]["genome_gff"]).resolve())

    annotation_table = data["input"].get("annotation_table")
    if annotation_table is not None:
        data["input"]["annotation_table"] = str((fixture_dir / annotation_table).resolve())

    domain_table = data["domains"].get("local_table")
    if domain_table is not None:
        data["domains"]["local_table"] = str((fixture_dir / domain_table).resolve())

    orthology_table = data["orthology"].get("local_table")
    if orthology_table is not None:
        data["orthology"]["local_table"] = str((fixture_dir / orthology_table).resolve())

    profile_table = data["phylogenetic_profile"].get("local_table")
    if profile_table is not None:
        data["phylogenetic_profile"]["local_table"] = str((fixture_dir / profile_table).resolve())

    interaction_table = data["known_interactions"].get("local_table")
    if interaction_table is not None:
        data["known_interactions"]["local_table"] = str((fixture_dir / interaction_table).resolve())

    fusion_table = data["fusion"].get("local_table")
    if fusion_table is not None:
        data["fusion"]["local_table"] = str((fixture_dir / fusion_table).resolve())

    domain_rules_path = data["domains"].get("rules_path")
    if domain_rules_path is not None:
        data["domains"]["rules_path"] = str((fixture_dir / domain_rules_path).resolve())

    rules_path = data["functional_complementarity"].get("rules_path")
    if rules_path is not None:
        data["functional_complementarity"]["rules_path"] = str((fixture_dir / rules_path).resolve())

    data["localization"]["enabled"] = True
    data["localization"]["source"] = "annotation_only"

    data["output"]["directory"] = str(tmp_path / "generated")

    target = tmp_path / "config.yaml"
    target.write_text(
        yaml.safe_dump(data, sort_keys=False),
        encoding="utf-8",
    )
    return target


def test_pipeline_gene_context_scores_and_jsonl_round_trip(
    valid_config_path: Path, tmp_path: Path
) -> None:
    result = InteractionCandidatePipeline().run(e2e_config(valid_config_path, tmp_path))
    assert len(result.bundles) == 13

    for bundle in result.bundles:
        assert len(bundle.genome_context) == 1
        assert bundle.engine_statuses["gene_context"] is bundle.genome_context[0].status

        assert len(bundle.operon) == 1
        assert bundle.engine_statuses["operon"] is bundle.operon[0].status

        assert len(bundle.functional) >= 1
        assert bundle.engine_statuses["functional_complementarity"] is bundle.functional[0].status
        assert len(bundle.domains) >= 1
        assert bundle.engine_statuses["domains"] is bundle.domains[0].status
        assert len(bundle.localization) == 1
        assert bundle.engine_statuses["localization"] is bundle.localization[0].status
        assert len(bundle.orthology) >= 1
        assert bundle.engine_statuses["orthology"] is bundle.orthology[0].status

        assert len(bundle.phylogenetic_profile) == 1
        assert (
            bundle.engine_statuses["phylogenetic_profile"] is bundle.phylogenetic_profile[0].status
        )

        assert len(bundle.fusion) == 1
        assert bundle.engine_statuses["fusion"] is bundle.fusion[0].status

        assert len(bundle.known_interactions) == 1
        assert bundle.engine_statuses["known_interactions"] is bundle.known_interactions[0].status

        assert all(
            status is EvidenceStatus.NOT_RUN
            for engine, status in bundle.engine_statuses.items()
            if engine
            not in {
                "gene_context",
                "operon",
                "domains",
                "functional_complementarity",
                "localization",
                "orthology",
                "phylogenetic_profile",
                "fusion",
                "known_interactions",
            }
        )

        scores = bundle.score.model_dump()
        assert all(value is None for key, value in scores.items() if key.endswith("_score"))
        assert bundle.score.contradiction_penalty is None
        assert bundle.score.evidence_completeness is None
        assert bundle.evidence_tier is None

    bundles_by_id = {bundle.candidate_id: bundle for bundle in result.bundles}

    near_localization = bundles_by_id["NEAR_001"].localization[0]
    assert near_localization.status is EvidenceStatus.AVAILABLE
    assert near_localization.query_compartment == "cytosolic"
    assert near_localization.candidate_compartment == "cytosolic"
    assert near_localization.compatibility is True
    assert near_localization.transmembrane_helices == 0
    assert near_localization.topology == "none"

    membrane_localization = bundles_by_id["MEM_001"].localization[0]
    assert membrane_localization.status is EvidenceStatus.AVAILABLE
    assert membrane_localization.query_compartment == "cytosolic"
    assert membrane_localization.candidate_compartment == "membrane"
    assert membrane_localization.compatibility is False
    assert "different_compartment" in (membrane_localization.conflicting_terms)
    assert membrane_localization.topology == "multi_pass"

    assert JsonlEvidenceBundleWriter().read(result.evidence_path) == result.bundles
    assert all(
        CandidateEvidenceBundle.model_validate_json(line)
        for line in result.evidence_path.read_text(encoding="utf-8").splitlines()
    )


def test_candidate_tsv_header_and_manifest_provenance(
    valid_config_path: Path,
    tmp_path: Path,
) -> None:
    config = e2e_config(valid_config_path, tmp_path)
    result = InteractionCandidatePipeline().run(config)

    with result.candidate_table_path.open(
        encoding="utf-8",
        newline="",
    ) as handle:
        reader = csv.reader(handle, delimiter="\t")
        assert tuple(next(reader)) == CANDIDATE_COLUMNS
        assert len(list(reader)) == 13

    with result.candidate_table_path.open(
        encoding="utf-8",
        newline="",
    ) as handle:
        rows_by_id = {
            row["candidate_id"]: row
            for row in csv.DictReader(
                handle,
                delimiter="\t",
            )
        }

    assert rows_by_id["NEAR_001"]["distance_bp"] == "29"
    assert rows_by_id["CONTIG2_001"]["distance_bp"] == ""
    assert rows_by_id["FRAG_001"]["gene_context_status"] == "missing"

    assert rows_by_id["NEAR_001"]["orthology_status"] == "available"
    assert rows_by_id["NEAR_001"]["orthology_pair_supported"] == "True"
    assert rows_by_id["NEAR_001"]["orthology_rule_version"] == ORTHOLOGY_ENGINE_VERSION

    assert rows_by_id["NEAR_001"]["phylogenetic_profile_status"] == "available"
    assert rows_by_id["NEAR_001"]["phylogenetic_profile_similarity"] == "1.0"
    assert rows_by_id["NEAR_001"]["phylogenetic_profile_pair_supported"] == "True"
    assert (
        rows_by_id["NEAR_001"]["phylogenetic_profile_rule_version"]
        == PHYLOGENETIC_PROFILE_ENGINE_VERSION
    )
    assert rows_by_id["PARA_001"]["phylogenetic_profile_status"] == "missing"
    assert rows_by_id["PARA_001"]["phylogenetic_profile_pair_supported"] == ""

    assert rows_by_id["NEAR_001"]["fusion_status"] == "available"
    assert rows_by_id["NEAR_001"]["fusion_supporting_record_count"] == "2"
    assert rows_by_id["NEAR_001"]["fusion_qualifying_record_count"] == "2"
    assert rows_by_id["NEAR_001"]["fusion_pair_supported"] == "True"
    assert rows_by_id["NEAR_001"]["fusion_rule_version"] == FUSION_ENGINE_VERSION
    assert rows_by_id["MEM_001"]["fusion_pair_supported"] == "False"
    assert rows_by_id["PARA_001"]["fusion_status"] == "missing"
    assert rows_by_id["PARA_001"]["fusion_pair_supported"] == ""

    assert rows_by_id["NEAR_001"]["known_interactions_status"] == "available"
    assert rows_by_id["NEAR_001"]["known_interactions_supporting_record_count"] == "2"
    assert rows_by_id["NEAR_001"]["known_interactions_direct_record_count"] == "2"
    assert rows_by_id["NEAR_001"]["known_interactions_pair_supported"] == "True"
    assert rows_by_id["MID_001"]["known_interactions_pair_supported"] == "False"
    assert rows_by_id["MID_001"]["known_interactions_functional_association_supported"] == "True"
    assert rows_by_id["PARA_001"]["known_interactions_status"] == "missing"
    assert rows_by_id["PARA_001"]["known_interactions_pair_supported"] == ""
    assert rows_by_id["FAR_001"]["known_interactions_status"] == "available"
    assert rows_by_id["FAR_001"]["known_interactions_pair_supported"] == "False"
    assert (
        rows_by_id["NEAR_001"]["known_interactions_rule_version"]
        == KNOWN_INTERACTIONS_ENGINE_VERSION
    )

    assert rows_by_id["NEAR_001"]["localization_status"] == "available"
    assert rows_by_id["NEAR_001"]["localization_query_compartment"] == "cytosolic"
    assert rows_by_id["NEAR_001"]["localization_compartment"] == "cytosolic"
    assert rows_by_id["NEAR_001"]["localization_compatibility"] == "True"
    assert rows_by_id["NEAR_001"]["localization_tm_helices"] == "0"
    assert rows_by_id["NEAR_001"]["localization_topology"] == "none"

    assert rows_by_id["MEM_001"]["localization_status"] == "available"
    assert rows_by_id["MEM_001"]["localization_compartment"] == "membrane"
    assert rows_by_id["MEM_001"]["localization_compatibility"] == "False"
    assert rows_by_id["MEM_001"]["localization_topology"] == "multi_pass"
    assert "different_compartment" in (rows_by_id["MEM_001"]["localization_conflicting_terms"])

    manifest = RunManifest.model_validate_json(
        result.manifest_path.read_text(
            encoding="utf-8",
        )
    )

    assert manifest.config_sha256
    assert all(item.sha256 for item in manifest.input_files)
    input_files_by_name = {item.logical_name: item for item in manifest.input_files}

    assert "orthology_local_table" in input_files_by_name
    assert input_files_by_name["orthology_local_table"].required is True
    assert input_files_by_name["orthology_local_table"].sha256
    assert input_files_by_name["domain_local_table"].required is True
    assert input_files_by_name["domain_local_table"].sha256
    assert input_files_by_name["domain_pair_rules"].required is True
    assert input_files_by_name["domain_pair_rules"].sha256
    assert manifest.package_version == "0.1.0"
    assert manifest.git_commit
    assert manifest.normalization_rule_version == NORMALIZATION_RULE_VERSION
    assert manifest.gene_context_rule_version == GENE_CONTEXT_RULE_VERSION
    assert manifest.orthology_rule_version == ORTHOLOGY_ENGINE_VERSION
    assert manifest.phylogenetic_profile_rule_version == PHYLOGENETIC_PROFILE_ENGINE_VERSION
    assert "phylogenetic_profile_local_table" in input_files_by_name
    assert input_files_by_name["phylogenetic_profile_local_table"].required is True
    assert input_files_by_name["phylogenetic_profile_local_table"].sha256
    assert manifest.fusion_rule_version == FUSION_ENGINE_VERSION
    assert "fusion_local_table" in input_files_by_name
    assert input_files_by_name["fusion_local_table"].required is True
    assert input_files_by_name["fusion_local_table"].sha256
    assert manifest.known_interactions_rule_version == KNOWN_INTERACTIONS_ENGINE_VERSION
    assert "known_interactions_local_table" in input_files_by_name
    assert input_files_by_name["known_interactions_local_table"].required is True
    assert input_files_by_name["known_interactions_local_table"].sha256
    assert manifest.policy_settings["fragment_policy"] == "flag"
    assert manifest.policy_settings["localization_rule_version"] == LOCALIZATION_ENGINE_VERSION
    assert "scoring_not_run" in manifest.incomplete_evidence_flags
    assert "gene_context_not_run" not in manifest.incomplete_evidence_flags
    assert "localization_not_run" not in manifest.incomplete_evidence_flags

    assert result.config_snapshot_path.is_file()
    assert result.warning_summary_path.is_file()
    assert "known_interaction_records_with_missing_confidence:1" in (
        result.warning_summary_path.read_text(encoding="utf-8")
    )
    assert "fusion_records_with_missing_component_coverage:1" in (
        result.warning_summary_path.read_text(encoding="utf-8")
    )
    assert result.excel_path is not None
    assert result.excel_path.is_file()

    workbook = load_workbook(
        result.excel_path,
        read_only=True,
        data_only=True,
    )

    assert "Orthology_Evidence" in workbook.sheetnames

    orthology_sheet = workbook["Orthology_Evidence"]
    orthology_headers = [
        cell.value
        for cell in next(
            orthology_sheet.iter_rows(
                min_row=1,
                max_row=1,
            )
        )
    ]
    orthology_rows = [
        dict(zip(orthology_headers, values, strict=True))
        for values in orthology_sheet.iter_rows(
            min_row=2,
            values_only=True,
        )
    ]
    orthology_by_candidate = {row["Candidate_ID"]: row for row in orthology_rows}

    assert orthology_by_candidate["NEAR_001"]["Status"] == "available"
    assert orthology_by_candidate["NEAR_001"]["Pair_Supported"] is True
    assert orthology_by_candidate["NEAR_001"]["Rule_Version"] == ORTHOLOGY_ENGINE_VERSION

    assert "Phylogenetic_Profile_Evidence" in workbook.sheetnames
    profile_sheet = workbook["Phylogenetic_Profile_Evidence"]
    profile_headers = [cell.value for cell in next(profile_sheet.iter_rows(max_row=1))]
    profile_rows = [
        dict(zip(profile_headers, values, strict=True))
        for values in profile_sheet.iter_rows(min_row=2, values_only=True)
    ]
    profiles_by_candidate = {row["Candidate_ID"]: row for row in profile_rows}
    assert profiles_by_candidate["NEAR_001"]["Profile_Similarity"] == 1.0
    assert profiles_by_candidate["NEAR_001"]["Pair_Supported"] is True
    assert profiles_by_candidate["PARA_001"]["Status"] == "missing"

    assert "Known_Interactions_Evidence" in workbook.sheetnames
    interaction_sheet = workbook["Known_Interactions_Evidence"]
    interaction_headers = [cell.value for cell in next(interaction_sheet.iter_rows(max_row=1))]
    interaction_rows = [
        dict(zip(interaction_headers, values, strict=True))
        for values in interaction_sheet.iter_rows(min_row=2, values_only=True)
    ]
    interactions_by_candidate = {row["Candidate_ID"]: row for row in interaction_rows}
    assert interactions_by_candidate["NEAR_001"]["Pair_Supported"] is True
    assert interactions_by_candidate["NEAR_001"]["Direct_Record_Count"] == 2
    assert interactions_by_candidate["NEAR_001"]["Independent_Publication_Count"] == 1
    assert interactions_by_candidate["NEAR_001"]["Independent_Source_Count"] == 2
    assert interactions_by_candidate["MID_001"]["Functional_Association_Supported"] is True
    assert interactions_by_candidate["PARA_001"]["Status"] == "missing"
    assert interactions_by_candidate["FAR_001"]["Pair_Supported"] is False

    assert "Localization_Evidence" in workbook.sheetnames

    worksheet = workbook["Localization_Evidence"]
    headers = [
        cell.value
        for cell in next(
            worksheet.iter_rows(
                min_row=1,
                max_row=1,
            )
        )
    ]
    rows = [
        dict(zip(headers, values, strict=True))
        for values in worksheet.iter_rows(
            min_row=2,
            values_only=True,
        )
    ]

    localization_rows = {row["Candidate_ID"]: row for row in rows}

    assert localization_rows["NEAR_001"]["Status"] == "available"
    assert localization_rows["NEAR_001"]["Query_Compartment"] == "cytosolic"
    assert localization_rows["NEAR_001"]["Candidate_Compartment"] == "cytosolic"
    assert localization_rows["NEAR_001"]["Compatibility"] is True
    assert localization_rows["NEAR_001"]["Topology"] == "none"

    assert localization_rows["MEM_001"]["Status"] == "available"
    assert localization_rows["MEM_001"]["Candidate_Compartment"] == "membrane"
    assert localization_rows["MEM_001"]["Compatibility"] is False
    assert localization_rows["MEM_001"]["Topology"] == "multi_pass"

    workbook.close()


def test_profile_is_shadow_only_and_disabled_is_not_run(
    valid_config_path: Path,
    tmp_path: Path,
) -> None:
    enabled_dir = tmp_path / "enabled"
    enabled_dir.mkdir()
    enabled_config = e2e_config(valid_config_path, enabled_dir)
    enabled_result = InteractionCandidatePipeline().run(enabled_config)

    disabled_data = yaml.safe_load(enabled_config.read_text(encoding="utf-8"))
    disabled_data["phylogenetic_profile"]["enabled"] = False
    disabled_data["output"]["directory"] = str(tmp_path / "disabled-output")
    disabled_config = tmp_path / "disabled.yaml"
    disabled_config.write_text(
        yaml.safe_dump(disabled_data, sort_keys=False),
        encoding="utf-8",
    )
    disabled_result = InteractionCandidatePipeline().run(disabled_config)

    assert [bundle.candidate_id for bundle in enabled_result.bundles] == [
        bundle.candidate_id for bundle in disabled_result.bundles
    ]
    for enabled, disabled in zip(
        enabled_result.bundles,
        disabled_result.bundles,
        strict=True,
    ):
        assert enabled.score == disabled.score
        assert enabled.evidence_tier == disabled.evidence_tier
        assert enabled.predicted_relationship_type == disabled.predicted_relationship_type
        assert disabled.phylogenetic_profile == []
        assert disabled.engine_statuses["phylogenetic_profile"] is EvidenceStatus.NOT_RUN


def test_fusion_is_shadow_only_in_enabled_disabled_ab_comparison(
    valid_config_path: Path,
    tmp_path: Path,
) -> None:
    enabled_dir = tmp_path / "fusion-enabled"
    enabled_dir.mkdir()
    enabled_config = e2e_config(valid_config_path, enabled_dir)
    enabled_result = InteractionCandidatePipeline().run(enabled_config)

    disabled_data = yaml.safe_load(enabled_config.read_text(encoding="utf-8"))
    disabled_data["fusion"]["enabled"] = False
    disabled_data["output"]["directory"] = str(tmp_path / "fusion-disabled-output")
    disabled_config = tmp_path / "fusion-disabled.yaml"
    disabled_config.write_text(yaml.safe_dump(disabled_data, sort_keys=False), encoding="utf-8")
    disabled_result = InteractionCandidatePipeline().run(disabled_config)

    omitted_data = dict(disabled_data)
    del omitted_data["fusion"]
    omitted_data["output"] = dict(omitted_data["output"])
    omitted_data["output"]["directory"] = str(tmp_path / "fusion-omitted-output")
    omitted_config = tmp_path / "fusion-omitted.yaml"
    omitted_config.write_text(yaml.safe_dump(omitted_data, sort_keys=False), encoding="utf-8")
    omitted_result = InteractionCandidatePipeline().run(omitted_config)
    assert all(not bundle.fusion for bundle in omitted_result.bundles)
    assert all(
        bundle.engine_statuses["fusion"] is EvidenceStatus.NOT_RUN
        for bundle in omitted_result.bundles
    )

    assert [bundle.candidate_id for bundle in enabled_result.bundles] == [
        bundle.candidate_id for bundle in disabled_result.bundles
    ]
    for enabled, disabled in zip(enabled_result.bundles, disabled_result.bundles, strict=True):
        assert enabled.candidate_disposition == disabled.candidate_disposition
        assert enabled.score == disabled.score
        assert enabled.predicted_relationship_type == disabled.predicted_relationship_type
        assert enabled.evidence_tier == disabled.evidence_tier
        assert enabled.orthology == disabled.orthology
        assert enabled.phylogenetic_profile == disabled.phylogenetic_profile
        assert disabled.fusion == []
        assert disabled.engine_statuses["fusion"] is EvidenceStatus.NOT_RUN


def test_known_interactions_are_shadow_only_in_enabled_disabled_ab_comparison(
    valid_config_path: Path,
    tmp_path: Path,
) -> None:
    enabled_dir = tmp_path / "known-enabled"
    enabled_dir.mkdir()
    enabled_config = e2e_config(valid_config_path, enabled_dir)
    enabled_result = InteractionCandidatePipeline().run(enabled_config)

    disabled_data = yaml.safe_load(enabled_config.read_text(encoding="utf-8"))
    disabled_data["known_interactions"]["enabled"] = False
    disabled_data["output"]["directory"] = str(tmp_path / "known-disabled-output")
    disabled_config = tmp_path / "known-disabled.yaml"
    disabled_config.write_text(yaml.safe_dump(disabled_data, sort_keys=False), encoding="utf-8")
    disabled_result = InteractionCandidatePipeline().run(disabled_config)

    omitted_data = dict(disabled_data)
    del omitted_data["known_interactions"]
    omitted_data["output"] = dict(omitted_data["output"])
    omitted_data["output"]["directory"] = str(tmp_path / "known-omitted-output")
    omitted_config = tmp_path / "known-omitted.yaml"
    omitted_config.write_text(yaml.safe_dump(omitted_data, sort_keys=False), encoding="utf-8")
    omitted_result = InteractionCandidatePipeline().run(omitted_config)

    assert [bundle.candidate_id for bundle in enabled_result.bundles] == [
        bundle.candidate_id for bundle in disabled_result.bundles
    ]
    for enabled, disabled in zip(enabled_result.bundles, disabled_result.bundles, strict=True):
        assert enabled.candidate_disposition == disabled.candidate_disposition
        assert enabled.score == disabled.score
        assert enabled.predicted_relationship_type == disabled.predicted_relationship_type
        assert enabled.evidence_tier == disabled.evidence_tier
        assert enabled.orthology == disabled.orthology
        assert enabled.phylogenetic_profile == disabled.phylogenetic_profile
        assert enabled.fusion == disabled.fusion
        assert disabled.known_interactions == []
        assert disabled.engine_statuses["known_interactions"] is EvidenceStatus.NOT_RUN

    assert all(not bundle.known_interactions for bundle in omitted_result.bundles)
    assert all(
        bundle.engine_statuses["known_interactions"] is EvidenceStatus.NOT_RUN
        for bundle in omitted_result.bundles
    )


def test_generate_candidates_cli_e2e_has_no_ranking_output(
    valid_config_path: Path, tmp_path: Path
) -> None:
    config = e2e_config(valid_config_path, tmp_path)
    result = CliRunner().invoke(app, ["generate-candidates", "--config", str(config)])
    assert result.exit_code == 0, result.stdout
    assert "query_count: 1" in result.stdout
    assert "protein_count: 13" in result.stdout
    assert "query_candidate_pair_count: 13" in result.stdout
    assert "duplicate_group_count: 1" in result.stdout
    assert "same_contig_pair_count:" in result.stdout
    assert "different_contig_pair_count: 1" in result.stdout
    assert "overlapping_pair_count:" in result.stdout
    assert "output_path:" in result.stdout
    assert "ranking" not in result.stdout.casefold()
    output = tmp_path / "generated"
    assert {path.name for path in output.iterdir()} == {
        "ProteinInteractionHunter.xlsx",
        "candidate_evidence_bundle.jsonl",
        "candidate_table.tsv",
        "config.snapshot.yaml",
        "run_manifest.json",
        "warning_summary.tsv",
    }
    first = (output / "candidate_table.tsv").read_text(encoding="utf-8")
    rerun = CliRunner().invoke(app, ["generate-candidates", "--config", str(config)])
    assert rerun.exit_code == 0
    assert (output / "candidate_table.tsv").read_text(encoding="utf-8") == first
    payload = json.loads((output / "run_manifest.json").read_text(encoding="utf-8"))
    assert payload["normalization_rule_version"] == NORMALIZATION_RULE_VERSION
    assert payload["gene_context_rule_version"] == GENE_CONTEXT_RULE_VERSION


def test_enabled_known_interactions_requires_local_table(
    valid_config_path: Path,
    tmp_path: Path,
) -> None:
    config = e2e_config(valid_config_path, tmp_path)
    data = yaml.safe_load(config.read_text(encoding="utf-8"))
    data["known_interactions"]["local_table"] = None
    config.write_text(yaml.safe_dump(data, sort_keys=False), encoding="utf-8")
    with pytest.raises(InputValidationError, match="known_interactions.local_table is required"):
        InteractionCandidatePipeline().run(config)


def test_integrated_scoring_outputs_and_enabled_disabled_ab(
    valid_config_path: Path,
    tmp_path: Path,
) -> None:
    disabled_dir = tmp_path / "scoring-disabled"
    disabled_dir.mkdir()
    disabled_config = e2e_config(valid_config_path, disabled_dir)
    disabled_result = InteractionCandidatePipeline().run(disabled_config)

    enabled_data = yaml.safe_load(disabled_config.read_text(encoding="utf-8"))
    enabled_data["scoring"]["enabled"] = True
    enabled_data["output"]["directory"] = str(tmp_path / "scoring-enabled-output")
    enabled_config = tmp_path / "scoring-enabled.yaml"
    enabled_config.write_text(yaml.safe_dump(enabled_data, sort_keys=False), encoding="utf-8")
    enabled_result = InteractionCandidatePipeline().run(enabled_config)

    assert [item.candidate_id for item in disabled_result.bundles] == [
        item.candidate_id for item in enabled_result.bundles
    ]
    for disabled, enabled in zip(disabled_result.bundles, enabled_result.bundles, strict=True):
        assert disabled.candidate_disposition == enabled.candidate_disposition
        assert disabled.predicted_relationship_type == enabled.predicted_relationship_type
        assert disabled.evidence_tier is None
        assert enabled.evidence_tier is None
        assert disabled.genome_context == enabled.genome_context
        assert disabled.operon == enabled.operon
        assert disabled.domains == enabled.domains
        assert disabled.functional == enabled.functional
        assert disabled.localization == enabled.localization
        assert disabled.orthology == enabled.orthology
        assert disabled.phylogenetic_profile == enabled.phylogenetic_profile
        assert disabled.fusion == enabled.fusion
        assert disabled.known_interactions == enabled.known_interactions
        assert disabled.integrated_scoring == []
        assert disabled.engine_statuses["scoring"] is EvidenceStatus.NOT_RUN
        assert all(
            value is None
            for key, value in disabled.score.model_dump().items()
            if key.endswith("_score")
        )
        assert len(enabled.integrated_scoring) == 1
        assert enabled.engine_statuses["scoring"] is enabled.integrated_scoring[0].status

    with enabled_result.candidate_table_path.open(encoding="utf-8", newline="") as handle:
        enabled_rows = list(csv.DictReader(handle, delimiter="	"))
    with disabled_result.candidate_table_path.open(encoding="utf-8", newline="") as handle:
        disabled_rows = list(csv.DictReader(handle, delimiter="	"))
    assert [row["candidate_id"] for row in enabled_rows] == [
        row["candidate_id"] for row in disabled_rows
    ]
    enabled_by_id = {row["candidate_id"]: row for row in enabled_rows}
    assert enabled_by_id["NEAR_001"]["integrated_score"]
    assert enabled_by_id["NEAR_001"]["rank"]
    assert enabled_by_id["NEAR_001"]["score_component_known_interactions"] == "1.0"
    assert enabled_by_id["MID_001"]["score_component_known_interactions"] == "0.5"
    assert enabled_by_id["QUERY_001"]["rank"] == ""
    assert all(row["integrated_score"] == "" for row in disabled_rows)
    assert all(row["rank"] == "" for row in disabled_rows)

    enabled_bundles = {item.candidate_id: item for item in enabled_result.bundles}
    near = enabled_bundles["NEAR_001"].integrated_scoring[0]
    mid = enabled_bundles["MID_001"].integrated_scoring[0]
    assert near.rank is not None
    assert mid.rank is not None
    assert near.output_score is not None
    assert mid.output_score is not None
    assert near.output_score > mid.output_score
    assert near.rank < mid.rank
    assert sum(
        component.weighted_contribution or 0.0
        for component in near.component_scores
        if component.applied
    ) == pytest.approx(near.raw_weighted_sum, abs=1e-12)

    manifest = RunManifest.model_validate_json(
        enabled_result.manifest_path.read_text(encoding="utf-8")
    )
    assert manifest.scoring_rule_version == SCORING_RULE_VERSION
    assert manifest.scoring_config_snapshot["enabled"] is True
    assert manifest.scoring_config_snapshot["weights"]["fusion"] == 1.5
    assert manifest.scoring_config_snapshot["missing_policy"] == "exclude_from_denominator"
    assert "scoring_not_run" not in manifest.incomplete_evidence_flags

    disabled_manifest = RunManifest.model_validate_json(
        disabled_result.manifest_path.read_text(encoding="utf-8")
    )
    assert disabled_manifest.scoring_rule_version is None
    assert disabled_manifest.scoring_config_snapshot == {}
    assert "scoring_not_run" in disabled_manifest.incomplete_evidence_flags

    assert enabled_result.excel_path is not None
    workbook = load_workbook(enabled_result.excel_path, read_only=True, data_only=True)
    assert "Integrated_Scoring" in workbook.sheetnames
    assert "Scoring_Components" in workbook.sheetnames
    integrated_sheet = workbook["Integrated_Scoring"]
    component_sheet = workbook["Scoring_Components"]
    assert integrated_sheet.max_row == len(enabled_result.bundles) + 1
    assert component_sheet.max_row == len(enabled_result.bundles) * 9 + 1
    workbook.close()


def test_omitted_scoring_section_is_disabled_in_pipeline(
    valid_config_path: Path,
    tmp_path: Path,
) -> None:
    config = e2e_config(valid_config_path, tmp_path)
    data = yaml.safe_load(config.read_text(encoding="utf-8"))
    del data["scoring"]
    config.write_text(yaml.safe_dump(data, sort_keys=False), encoding="utf-8")
    result = InteractionCandidatePipeline().run(config)
    assert all(not bundle.integrated_scoring for bundle in result.bundles)
    assert all(
        bundle.engine_statuses["scoring"] is EvidenceStatus.NOT_RUN for bundle in result.bundles
    )
    assert all(bundle.evidence_tier is None for bundle in result.bundles)


def test_scoring_insufficiency_is_warned_and_not_ranked(
    valid_config_path: Path,
    tmp_path: Path,
) -> None:
    config = e2e_config(valid_config_path, tmp_path)
    data = yaml.safe_load(config.read_text(encoding="utf-8"))
    data["scoring"]["enabled"] = True
    data["scoring"]["minimum_evidence_categories"] = 6
    config.write_text(yaml.safe_dump(data, sort_keys=False), encoding="utf-8")
    result = InteractionCandidatePipeline().run(config)
    assert all(not bundle.integrated_scoring[0].sufficient_evidence for bundle in result.bundles)
    assert all(bundle.integrated_scoring[0].rank is None for bundle in result.bundles)
    assert "insufficient_evidence_for_formal_score" in (
        result.warning_summary_path.read_text(encoding="utf-8")
    )
