"""MVP-1 final integration audit across all local evidence engines."""

from __future__ import annotations

import csv
import json
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, cast

import pytest
import yaml
from jsonschema import Draft202012Validator  # type: ignore[import-untyped]
from openpyxl import load_workbook
from typer.testing import CliRunner

from protein_interaction_hunter.adapters.local.domain_rules import LocalDomainRulesLoader
from protein_interaction_hunter.adapters.local.domains import LocalDomainTsvLoader
from protein_interaction_hunter.adapters.local.orthology import LocalOrthologyTsvLoader
from protein_interaction_hunter.application.domain_pairs import (
    build_domain_index,
    evaluate_domain_pairs,
)
from protein_interaction_hunter.application.orthology import (
    build_orthology_index,
    evaluate_orthology_pair,
)
from protein_interaction_hunter.application.pipeline import (
    _UNIMPLEMENTED_ENGINES,
    InteractionCandidatePipeline,
    PipelineResult,
)
from protein_interaction_hunter.cli import app
from protein_interaction_hunter.config import AppConfig, load_config
from protein_interaction_hunter.exceptions import ProteinInteractionHunterError
from protein_interaction_hunter.manifest import build_input_file_manifest
from protein_interaction_hunter.models.enums import (
    CandidateDisposition,
    EvidenceStatus,
    EvidenceTier,
)
from protein_interaction_hunter.models.evidence import CandidateEvidenceBundle
from protein_interaction_hunter.models.run import RunManifest
from protein_interaction_hunter.outputs.excel import EXCEL_SHEETS

FIXTURES = Path(__file__).resolve().parents[1] / "fixtures"
ALL_ENGINES_CONFIG = FIXTURES / "config.all_engines.yaml"
RAW_FIELDS = (
    "genome_context",
    "operon",
    "domains",
    "functional",
    "localization",
    "orthology",
    "phylogenetic_profile",
    "fusion",
    "known_interactions",
)
ENGINE_FIELDS = {
    "gene_context": ("genome_context", "operon"),
    "domains": ("domains",),
    "functional_complementarity": ("functional",),
    "localization": ("localization",),
    "orthology": ("orthology",),
    "phylogenetic_profile": ("phylogenetic_profile",),
    "fusion": ("fusion",),
    "known_interactions": ("known_interactions",),
}
ENGINE_COMPONENTS = {
    "gene_context": ("genome_context", "operon_proxy"),
    "domains": ("domain_pair",),
    "functional_complementarity": ("functional_complementarity",),
    "localization": ("localization",),
    "orthology": ("orthology",),
    "phylogenetic_profile": ("phylogenetic_profile",),
    "fusion": ("fusion",),
    "known_interactions": ("known_interactions",),
}


def _absolute_config_data(output: Path) -> dict[str, Any]:
    data = cast(dict[str, Any], yaml.safe_load(ALL_ENGINES_CONFIG.read_text(encoding="utf-8")))
    for key in ("proteome_fasta", "genome_gff", "annotation_table"):
        value = data["input"].get(key)
        if value:
            data["input"][key] = str((FIXTURES / value).resolve())
    for section in (
        "orthology",
        "phylogenetic_profile",
        "fusion",
        "domains",
        "known_interactions",
    ):
        value = data[section].get("local_table")
        if value:
            data[section]["local_table"] = str((FIXTURES / value).resolve())
    for section in ("domains", "functional_complementarity"):
        value = data[section].get("rules_path")
        if value:
            data[section]["rules_path"] = str((FIXTURES / value).resolve())
    data["output"]["directory"] = str(output.resolve())
    data["cache"]["directory"] = str((output.parent / ".cache").resolve())
    data["logging"]["directory"] = str((output.parent / "logs").resolve())
    return data


def _write_config(path: Path, data: dict[str, Any]) -> Path:
    path.write_text(yaml.safe_dump(data, sort_keys=False), encoding="utf-8", newline="\n")
    return path


def _run_config(directory: Path, mutate: Any | None = None) -> tuple[PipelineResult, AppConfig]:
    directory.mkdir(parents=True, exist_ok=True)
    data = _absolute_config_data(directory / "output")
    if mutate is not None:
        mutate(data)
    path = _write_config(directory / "config.yaml", data)
    return InteractionCandidatePipeline().run(path), load_config(path)


@pytest.fixture(scope="module")
def canonical_audit(tmp_path_factory: pytest.TempPathFactory) -> tuple[PipelineResult, AppConfig]:
    return _run_config(tmp_path_factory.mktemp("mvp1-final-canonical"))


def _bundles_by_pair(
    result: PipelineResult,
) -> dict[tuple[str, str], CandidateEvidenceBundle]:
    return {(item.query_id, item.candidate_id): item for item in result.bundles}


def _workbook_values(path: Path) -> dict[str, list[tuple[Any, ...]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    values = {
        name: list(workbook[name].iter_rows(values_only=True)) for name in workbook.sheetnames
    }
    workbook.close()
    return values


def _normalized_manifest(path: Path) -> dict[str, Any]:
    payload = cast(dict[str, Any], json.loads(path.read_text(encoding="utf-8")))
    payload.pop("started_at")
    payload.pop("completed_at")
    return payload


def test_all_engine_canonical_e2e(
    canonical_audit: tuple[PipelineResult, AppConfig],
) -> None:
    result, _ = canonical_audit
    assert _UNIMPLEMENTED_ENGINES == ()
    assert result.summary.query_count == 1
    assert result.summary.pair_count == 13
    assert len(result.bundles) == 13
    assert [item.candidate_id for item in result.bundles] == sorted(
        item.candidate_id for item in result.bundles
    )
    for bundle in result.bundles:
        for engine in (
            "gene_context",
            "operon",
            "domains",
            "functional_complementarity",
            "localization",
            "orthology",
            "phylogenetic_profile",
            "fusion",
            "known_interactions",
            "scoring",
            "evidence_tiers",
        ):
            assert bundle.engine_statuses[engine] is not EvidenceStatus.NOT_RUN
        assert len(bundle.integrated_scoring) == 1
        assert len(bundle.evidence_tiers) == 1
        score = bundle.integrated_scoring[0]
        tier = bundle.evidence_tiers[0]
        if (
            score.sufficient_evidence
            and bundle.candidate_disposition is not CandidateDisposition.EXCLUDED
        ):
            assert score.output_score is not None
            assert score.rank is not None
        if tier.tier_eligible:
            assert bundle.evidence_tier is not None

    self_pair = _bundles_by_pair(result)[("QUERY_001", "QUERY_001")]
    assert self_pair.candidate_disposition is CandidateDisposition.EXCLUDED
    assert self_pair.integrated_scoring[0].rank is None
    assert self_pair.evidence_tier is EvidenceTier.UNCLASSIFIED
    assert self_pair.evidence_tiers[0].failed_requirements == ["self_pair"]


def test_score_full_recomputation(
    canonical_audit: tuple[PipelineResult, AppConfig],
) -> None:
    result, config = canonical_audit
    for bundle in result.bundles:
        score = bundle.integrated_scoring[0]
        applied = [item for item in score.component_scores if item.applied]
        for component in applied:
            assert component.weighted_contribution == pytest.approx(
                (component.normalized_value or 0.0) * component.effective_weight,
                abs=1e-12,
            )
        raw = round(sum(item.weighted_contribution or 0.0 for item in applied), 12)
        weight = round(sum(item.effective_weight for item in applied), 12)
        normalized = round(max(0.0, min(1.0, raw / weight)), 12) if weight else None
        provisional = (
            round(normalized * config.scoring.output_scale, 12) if normalized is not None else None
        )
        assert score.raw_weighted_sum == raw
        assert score.available_weight == weight
        assert score.normalized_score == normalized
        assert score.provisional_score == provisional
        assert score.output_score == (provisional if score.sufficient_evidence else None)


def test_rank_full_recomputation(
    canonical_audit: tuple[PipelineResult, AppConfig],
) -> None:
    result, config = canonical_audit
    by_query: dict[str, list[CandidateEvidenceBundle]] = defaultdict(list)
    for bundle in result.bundles:
        by_query[bundle.query_id].append(bundle)
    for bundles in by_query.values():
        eligible = [
            item
            for item in bundles
            if item.candidate_disposition is not CandidateDisposition.EXCLUDED
            and item.integrated_scoring[0].sufficient_evidence
            and item.integrated_scoring[0].output_score is not None
        ]
        eligible.sort(
            key=lambda item: (
                -round(
                    item.integrated_scoring[0].output_score or 0.0,
                    config.scoring.tie_precision,
                ),
                -item.integrated_scoring[0].evidence_category_count,
                -item.integrated_scoring[0].available_weight,
                item.candidate_id,
            )
        )
        dense_rank = 0
        previous: float | None = None
        expected: dict[str, int] = {}
        for bundle in eligible:
            value = round(
                bundle.integrated_scoring[0].output_score or 0.0,
                config.scoring.tie_precision,
            )
            if previous is None or value != previous:
                dense_rank += 1
                previous = value
            expected[bundle.candidate_id] = dense_rank
        for bundle in bundles:
            assert bundle.integrated_scoring[0].rank == expected.get(bundle.candidate_id)


def _independent_tier(
    bundle: CandidateEvidenceBundle,
    config: AppConfig,
) -> tuple[EvidenceTier, EvidenceTier, list[str], str | None]:
    score = bundle.integrated_scoring[0]
    candidate = bundle.candidate
    assert candidate is not None
    known = bundle.known_interactions[0] if bundle.known_interactions else None
    fusion = bundle.fusion[0] if bundle.fusion else None
    direct = bool(
        known and known.status is EvidenceStatus.AVAILABLE and known.direct_interaction_supported
    )
    physical = bool(
        known and known.status is EvidenceStatus.AVAILABLE and known.physical_interaction_supported
    )
    functional = bool(
        known
        and known.status is EvidenceStatus.AVAILABLE
        and known.functional_association_supported
    )
    fusion_supported = bool(
        fusion and fusion.status is EvidenceStatus.AVAILABLE and fusion.pair_supported
    )
    high_specificity = int(direct or physical) + int(fusion_supported)
    ineligible: str | None = None
    if score.status is not EvidenceStatus.AVAILABLE:
        ineligible = f"scoring_status_{score.status.value}"
    elif candidate.protein_id == score.query_protein_id:
        ineligible = "self_pair"
    elif candidate.disposition is CandidateDisposition.EXCLUDED:
        ineligible = "candidate_excluded"
    elif not score.sufficient_evidence:
        ineligible = "insufficient_evidence"
    elif score.output_score is None:
        ineligible = "formal_score_missing"
    elif score.rank is None:
        ineligible = "rank_missing"
    if ineligible:
        return EvidenceTier.UNCLASSIFIED, EvidenceTier.UNCLASSIFIED, [], ineligible

    base = EvidenceTier.UNCLASSIFIED
    for name in ("tier_1", "tier_2", "tier_3", "tier_4"):
        threshold = getattr(config.evidence_tiers, name)
        if all(
            (
                (score.output_score or 0.0) >= threshold.minimum_score,
                score.evidence_category_count >= threshold.minimum_categories,
                score.evidence_component_count >= threshold.minimum_components,
                score.available_weight >= threshold.minimum_available_weight,
                score.negative_component_count <= threshold.maximum_negative_components,
                high_specificity >= threshold.minimum_high_specificity_components,
                not threshold.require_high_specificity_evidence or high_specificity > 0,
            )
        ):
            base = EvidenceTier(name)
            break
    if base is EvidenceTier.UNCLASSIFIED:
        return base, base, [], None

    priority = {
        EvidenceTier.TIER_1: 1,
        EvidenceTier.TIER_2: 2,
        EvidenceTier.TIER_3: 3,
        EvidenceTier.TIER_4: 4,
        EvidenceTier.UNCLASSIFIED: 5,
    }
    types = (
        set(known.interaction_types)
        if known and known.status is EvidenceStatus.AVAILABLE
        else set()
    )
    conditions = (
        (
            "explicit_conflict",
            score.negative_component_count > 0,
            config.evidence_tiers.explicit_conflict_tier_cap,
        ),
        (
            "predicted_only",
            bool(types) and types <= {"predicted"} and high_specificity == 0,
            config.evidence_tiers.predicted_only_tier_cap,
        ),
        (
            "functional_association_only",
            functional and not (direct or physical or fusion_supported),
            config.evidence_tiers.functional_association_only_tier_cap,
        ),
    )
    assigned: EvidenceTier = base
    caps: list[str] = []
    for cap_name, applies, cap_value in conditions:
        cap = EvidenceTier(cap_value)
        if applies and priority[cap] > priority[base]:
            caps.append(f"{cap_name}:{cap.value}")
        if applies and priority[cap] > priority[assigned]:
            assigned = cap
    return base, assigned, caps, None


def test_tier_full_recomputation(
    canonical_audit: tuple[PipelineResult, AppConfig],
) -> None:
    result, config = canonical_audit
    for bundle in result.bundles:
        expected_base, expected_assigned, expected_caps, reason = _independent_tier(bundle, config)
        tier = bundle.evidence_tiers[0]
        assert tier.base_tier is expected_base
        assert tier.assigned_tier is expected_assigned
        assert tier.applied_tier_caps == expected_caps
        if reason:
            assert tier.failed_requirements == [reason]


def test_cross_output_consistency(
    canonical_audit: tuple[PipelineResult, AppConfig],
) -> None:
    result, _ = canonical_audit
    assert result.excel_path is not None
    jsonl = {
        (item.query_id, item.candidate_id): item
        for item in (
            CandidateEvidenceBundle.model_validate_json(line)
            for line in result.evidence_path.read_text(encoding="utf-8").splitlines()
        )
    }
    with result.candidate_table_path.open(encoding="utf-8", newline="") as handle:
        tsv = {
            (row["query_id"], row["candidate_id"]): row
            for row in csv.DictReader(handle, delimiter="\t")
        }
    workbook = _workbook_values(result.excel_path)
    ranking_header = list(workbook["Candidate_Ranking"][0])
    ranking = {
        (row[ranking_header.index("Query_ID")], row[ranking_header.index("Candidate_ID")]): row
        for row in workbook["Candidate_Ranking"][1:]
    }
    scoring_header = list(workbook["Integrated_Scoring"][0])
    scoring = {
        (row[scoring_header.index("Query_ID")], row[scoring_header.index("Candidate_ID")]): row
        for row in workbook["Integrated_Scoring"][1:]
    }
    tier_header = list(workbook["Evidence_Tiers"][0])
    tiers = {
        (row[tier_header.index("Query_ID")], row[tier_header.index("Candidate_ID")]): row
        for row in workbook["Evidence_Tiers"][1:]
    }
    for bundle in result.bundles:
        pair = (bundle.query_id, bundle.candidate_id)
        assert jsonl[pair] == bundle
        score = bundle.integrated_scoring[0]
        tier = bundle.evidence_tiers[0]
        assert tsv[pair]["candidate_disposition"] == bundle.candidate_disposition.value
        assert float(tsv[pair]["integrated_score"]) == score.output_score
        assert (
            int(tsv[pair]["rank"]) == score.rank
            if score.rank is not None
            else tsv[pair]["rank"] == ""
        )
        assert tsv[pair]["evidence_tier"] == tier.assigned_tier.value
        assert ranking[pair][ranking_header.index("Total_Ranking_Score")] == score.output_score
        assert ranking[pair][ranking_header.index("Evidence_Tier")] == tier.assigned_tier.value
        assert scoring[pair][scoring_header.index("Integrated_Score")] == score.output_score
        assert tiers[pair][tier_header.index("Assigned_Tier")] == tier.assigned_tier.value

    summary_header = list(workbook["Tier_Summary"][0])
    summary = workbook["Tier_Summary"][1]
    assigned = Counter(item.evidence_tier.value for item in result.bundles if item.evidence_tier)
    assert summary[summary_header.index("Tier_1_Count")] == assigned["tier_1"]
    assert summary[summary_header.index("Tier_2_Count")] == assigned["tier_2"]
    assert summary[summary_header.index("Tier_3_Count")] == assigned["tier_3"]
    assert summary[summary_header.index("Tier_4_Count")] == assigned["tier_4"]
    assert summary[summary_header.index("Unclassified_Count")] == assigned["unclassified"]


def test_deterministic_three_runs(
    canonical_audit: tuple[PipelineResult, AppConfig],
) -> None:
    first, config = canonical_audit
    snapshots: list[dict[str, Any]] = []
    for result in (
        first,
        InteractionCandidatePipeline().run(config.output.directory.parent / "config.yaml"),
        InteractionCandidatePipeline().run(config.output.directory.parent / "config.yaml"),
    ):
        assert result.excel_path is not None
        snapshots.append(
            {
                "tsv": result.candidate_table_path.read_bytes(),
                "jsonl": result.evidence_path.read_bytes(),
                "warnings": result.warning_summary_path.read_bytes(),
                "excel": _workbook_values(result.excel_path),
                "manifest": _normalized_manifest(result.manifest_path),
            }
        )
    assert snapshots[0] == snapshots[1] == snapshots[2]


def test_disabled_mode_compatibility(tmp_path: Path) -> None:
    all_enabled, _ = _run_config(tmp_path / "all")

    def disable_tiers(data: dict[str, Any]) -> None:
        data["evidence_tiers"]["enabled"] = False

    tiers_off, _ = _run_config(tmp_path / "tiers-off", disable_tiers)
    for enabled, disabled in zip(all_enabled.bundles, tiers_off.bundles, strict=True):
        assert enabled.integrated_scoring == disabled.integrated_scoring
        assert disabled.evidence_tier is None
        assert disabled.evidence_tiers == []
        assert disabled.engine_statuses["evidence_tiers"] is EvidenceStatus.NOT_RUN
        for field in RAW_FIELDS:
            assert getattr(enabled, field) == getattr(disabled, field)

    def disable_scoring(data: dict[str, Any]) -> None:
        data["scoring"]["enabled"] = False
        data["evidence_tiers"]["enabled"] = False

    scoring_off, _ = _run_config(tmp_path / "scoring-off", disable_scoring)
    for enabled, disabled in zip(all_enabled.bundles, scoring_off.bundles, strict=True):
        for field in RAW_FIELDS:
            assert getattr(enabled, field) == getattr(disabled, field)
        assert disabled.integrated_scoring == []
        assert disabled.score.total_ranking_score is None
        assert disabled.evidence_tier is None

    def disable_all(data: dict[str, Any]) -> None:
        for section in (
            "gene_context",
            "orthology",
            "phylogenetic_profile",
            "fusion",
            "domains",
            "functional_complementarity",
            "localization",
            "known_interactions",
            "scoring",
            "evidence_tiers",
        ):
            data[section]["enabled"] = False

    all_off, _ = _run_config(tmp_path / "all-off", disable_all)
    assert len(all_off.bundles) == 13
    for bundle in all_off.bundles:
        assert all(not getattr(bundle, field) for field in RAW_FIELDS)
        assert bundle.integrated_scoring == []
        assert bundle.evidence_tier is None
        assert all(status is EvidenceStatus.NOT_RUN for status in bundle.engine_statuses.values())


@pytest.mark.parametrize("engine", tuple(ENGINE_FIELDS))
def test_individual_engine_ablation(
    canonical_audit: tuple[PipelineResult, AppConfig],
    tmp_path: Path,
    engine: str,
) -> None:
    canonical, _ = canonical_audit

    def ablate(data: dict[str, Any]) -> None:
        data[engine]["enabled"] = False

    ablated, _ = _run_config(tmp_path / engine, ablate)
    canonical_by_pair = _bundles_by_pair(canonical)
    for bundle in ablated.bundles:
        original = canonical_by_pair[(bundle.query_id, bundle.candidate_id)]
        assert bundle.candidate == original.candidate
        assert bundle.candidate_disposition == original.candidate_disposition
        assert bundle.predicted_relationship_type == original.predicted_relationship_type
        for field in RAW_FIELDS:
            if field not in ENGINE_FIELDS[engine]:
                assert getattr(bundle, field) == getattr(original, field)
            else:
                assert not getattr(bundle, field)
        score = bundle.integrated_scoring[0]
        components = {item.component_name: item for item in score.component_scores}
        for name in ENGINE_COMPONENTS[engine]:
            assert components[name].applied is False
            assert components[name].evidence_status is EvidenceStatus.NOT_RUN
            assert components[name].exclusion_reason == "evidence_not_run"
        assert score.available_weight == round(
            sum(item.effective_weight for item in score.component_scores if item.applied),
            12,
        )


def _reverse_tsv(source: Path, target: Path) -> None:
    lines = source.read_text(encoding="utf-8").splitlines()
    target.write_text("\n".join([lines[0], *reversed(lines[1:])]) + "\n", encoding="utf-8")


def _semantic_bundle(bundle: CandidateEvidenceBundle) -> dict[str, Any]:
    payload = bundle.model_dump(mode="json")
    payload.pop("run_id")
    return payload


def test_input_row_order_independence(
    canonical_audit: tuple[PipelineResult, AppConfig],
    tmp_path: Path,
) -> None:
    canonical, _ = canonical_audit
    data = _absolute_config_data(tmp_path / "reordered-output")
    sources = {
        "annotation_table": (data["input"], "annotation_table"),
        "domains": (data["domains"], "local_table"),
        "orthology": (data["orthology"], "local_table"),
        "phylogenetic_profile": (data["phylogenetic_profile"], "local_table"),
        "fusion": (data["fusion"], "local_table"),
        "known_interactions": (data["known_interactions"], "local_table"),
    }
    for name, (section, key) in sources.items():
        target = tmp_path / f"{name}.tsv"
        _reverse_tsv(Path(section[key]), target)
        section[key] = str(target)
    config_path = _write_config(tmp_path / "reordered.yaml", data)
    reordered = InteractionCandidatePipeline().run(config_path)
    assert [item.candidate_id for item in canonical.bundles] == [
        item.candidate_id for item in reordered.bundles
    ]
    for before, after in zip(canonical.bundles, reordered.bundles, strict=True):
        assert _semantic_bundle(before) == _semantic_bundle(after)


@pytest.mark.parametrize(
    ("section", "field", "expected_status"),
    (
        ("domains", "domains", EvidenceStatus.MISSING),
        ("orthology", "orthology", EvidenceStatus.MISSING),
        (
            "phylogenetic_profile",
            "phylogenetic_profile",
            EvidenceStatus.MISSING,
        ),
        ("fusion", "fusion", EvidenceStatus.AVAILABLE),
        ("known_interactions", "known_interactions", EvidenceStatus.AVAILABLE),
    ),
)
def test_header_only_optional_input_has_consistent_semantic_status(
    tmp_path: Path,
    section: str,
    field: str,
    expected_status: EvidenceStatus,
) -> None:
    data = _absolute_config_data(tmp_path / f"{section}-empty-output")
    source = Path(data[section]["local_table"])
    header = source.read_text(encoding="utf-8").splitlines()[0]
    empty = tmp_path / f"{section}-empty.tsv"
    empty.write_text(header + "\n", encoding="utf-8")
    data[section]["local_table"] = str(empty)
    result = InteractionCandidatePipeline().run(
        _write_config(tmp_path / f"{section}-empty.yaml", data)
    )
    for bundle in result.bundles:
        records = getattr(bundle, field)
        assert len(records) == 1
        assert records[0].status is expected_status
        assert bundle.engine_statuses[section] is expected_status
        component_name = ENGINE_COMPONENTS[section][0]
        component = {
            item.component_name: item for item in bundle.integrated_scoring[0].component_scores
        }[component_name]
        if expected_status is EvidenceStatus.MISSING:
            assert component.applied is False
            assert component.exclusion_reason == "evidence_missing"
        else:
            assert component.normalized_value == 0.0


@pytest.mark.parametrize(
    ("section", "key"),
    (
        ("domains", "local_table"),
        ("orthology", "local_table"),
        ("phylogenetic_profile", "local_table"),
        ("fusion", "local_table"),
        ("known_interactions", "local_table"),
    ),
)
def test_duplicate_record_resistance(
    canonical_audit: tuple[PipelineResult, AppConfig],
    tmp_path: Path,
    section: str,
    key: str,
) -> None:
    canonical, _ = canonical_audit
    data = _absolute_config_data(tmp_path / f"{section}-output")
    source = Path(data[section][key])
    lines = source.read_text(encoding="utf-8").splitlines()
    duplicate = tmp_path / f"{section}-duplicate.tsv"
    duplicate.write_text("\n".join([*lines, lines[-1]]) + "\n", encoding="utf-8")
    data[section][key] = str(duplicate)
    try:
        result = InteractionCandidatePipeline().run(
            _write_config(tmp_path / f"{section}.yaml", data)
        )
    except ProteinInteractionHunterError as exc:
        assert "duplicate" in str(exc).casefold()
        return
    canonical_by_pair = _bundles_by_pair(canonical)
    for bundle in result.bundles:
        original = canonical_by_pair[(bundle.query_id, bundle.candidate_id)]
        assert bundle.integrated_scoring == original.integrated_scoring
        assert bundle.evidence_tier == original.evidence_tier
        assert (
            bundle.evidence_tiers[0].high_specificity_component_count
            == original.evidence_tiers[0].high_specificity_component_count
        )
        if bundle.known_interactions:
            before = original.known_interactions[0]
            after = bundle.known_interactions[0]
            assert after.independent_publication_count == before.independent_publication_count
            assert after.independent_source_count == before.independent_source_count
        for component in bundle.integrated_scoring[0].component_scores:
            assert len(component.support_terms) == len(set(component.support_terms))


def test_domain_and_orthology_pair_semantics_are_symmetric() -> None:
    domain_index = build_domain_index(
        LocalDomainTsvLoader().load(FIXTURES / "synthetic_domains.tsv")
    )
    rules_path = FIXTURES / "rules/domain_pairs.v1.yaml"
    rules = LocalDomainRulesLoader().load(rules_path)
    forward_domain = evaluate_domain_pairs("QUERY_001", "NEAR_001", domain_index, rules, rules_path)
    reverse_domain = evaluate_domain_pairs("NEAR_001", "QUERY_001", domain_index, rules, rules_path)
    assert forward_domain[0].pair_matched is True
    assert reverse_domain[0].pair_matched is True
    assert forward_domain[0].pair_rule_id == reverse_domain[0].pair_rule_id
    forward_semantic_terms = {term for term in forward_domain[0].support_terms if ":" not in term}
    reverse_semantic_terms = {term for term in reverse_domain[0].support_terms if ":" not in term}
    assert forward_semantic_terms == reverse_semantic_terms

    orthology_index = build_orthology_index(
        LocalOrthologyTsvLoader().load(FIXTURES / "synthetic_orthology.tsv")
    )
    forward_orthology = evaluate_orthology_pair("QUERY_001", "NEAR_001", orthology_index)
    reverse_orthology = evaluate_orthology_pair("NEAR_001", "QUERY_001", orthology_index)
    assert forward_orthology[0].pair_supported is True
    assert reverse_orthology[0].pair_supported is True
    assert forward_orthology[0].support_terms == reverse_orthology[0].support_terms


def test_manifest_hash_stability_and_sensitivity(tmp_path: Path) -> None:
    path = tmp_path / "input.tsv"
    path.write_text("a\tb\n1\t2\n", encoding="utf-8")
    first = build_input_file_manifest("audit", path, required=True)
    second = build_input_file_manifest("audit", path, required=True)
    assert first.sha256 == second.sha256
    assert first.size_bytes == second.size_bytes
    path.write_text("a\tb\n1\t3\n", encoding="utf-8")
    changed = build_input_file_manifest("audit", path, required=True)
    assert changed.sha256 != first.sha256
    assert changed.size_bytes == first.size_bytes


def test_schema_and_manifest_warning_audit(
    canonical_audit: tuple[PipelineResult, AppConfig],
) -> None:
    result, _ = canonical_audit
    root = Path(__file__).resolve().parents[2]
    bundle_schema = json.loads(
        (root / "schemas/candidate_evidence_bundle.schema.json").read_text(encoding="utf-8")
    )
    manifest_schema = json.loads(
        (root / "schemas/run_manifest.schema.json").read_text(encoding="utf-8")
    )
    bundle_validator = Draft202012Validator(bundle_schema)
    for line in result.evidence_path.read_text(encoding="utf-8").splitlines():
        bundle_validator.validate(json.loads(line))
    manifest_payload = json.loads(result.manifest_path.read_text(encoding="utf-8"))
    Draft202012Validator(manifest_schema).validate(manifest_payload)
    manifest = RunManifest.model_validate(manifest_payload)
    assert manifest.schema_version == "1.3"
    assert manifest.scoring_config_snapshot["enabled"] is True
    assert manifest.evidence_tier_config_snapshot["enabled"] is True
    assert {item.logical_name for item in manifest.input_files}.isdisjoint(
        {
            "candidate_table",
            "candidate_evidence_bundle",
            "ProteinInteractionHunter.xlsx",
        }
    )
    with result.warning_summary_path.open(encoding="utf-8", newline="") as handle:
        warnings = {
            row["warning"]: int(row["count"]) for row in csv.DictReader(handle, delimiter="\t")
        }
    expected = Counter(warning for bundle in result.bundles for warning in bundle.warnings)
    assert warnings == expected
    assert set(manifest.warnings) == set(expected)
    assert not any(item.startswith("evidence_tier_unclassified:") for item in warnings)


def test_excel_contract_and_output_files(
    canonical_audit: tuple[PipelineResult, AppConfig],
) -> None:
    result, _ = canonical_audit
    assert result.excel_path is not None
    workbook = _workbook_values(result.excel_path)
    assert list(workbook) == list(EXCEL_SHEETS)
    assert len(workbook["Candidate_Ranking"]) == 14
    assert len(workbook["Integrated_Scoring"]) == 14
    assert len(workbook["Evidence_Tiers"]) == 14
    assert len(workbook["Scoring_Components"]) == 13 * 9 + 1
    assert len(workbook["Tier_Summary"]) == 2
    assert {path.name for path in result.summary.output_path.iterdir() if path.is_file()} == {
        "ProteinInteractionHunter.xlsx",
        "candidate_evidence_bundle.jsonl",
        "candidate_table.tsv",
        "config.snapshot.yaml",
        "run_manifest.json",
        "warning_summary.tsv",
    }


def test_cli_text_reflects_mvp1_and_tier_support() -> None:
    runner = CliRunner()
    top = runner.invoke(app, ["--help"])
    command = runner.invoke(app, ["generate-candidates", "--help"])
    assert top.exit_code == 0
    assert command.exit_code == 0
    combined = f"{top.stdout}\n{command.stdout}"
    assert "MVP-1B" not in combined
    assert "MVP-0" not in combined
    assert "tiers remain unset" not in combined.casefold()
    assert "scoring" in combined.casefold()
    assert "tier" in combined.casefold()


def test_readme_reflects_current_mvp1_capabilities() -> None:
    readme = (Path(__file__).resolve().parents[2] / "README.md").read_text(encoding="utf-8")
    assert "Current status: MVP-1B" not in readme
    assert "does **not** infer operons" not in readme
    assert "It does not calculate or print ranks, scores" not in readme
    assert "integrated scoring" in readme.casefold()
    assert "evidence tier" in readme.casefold()
